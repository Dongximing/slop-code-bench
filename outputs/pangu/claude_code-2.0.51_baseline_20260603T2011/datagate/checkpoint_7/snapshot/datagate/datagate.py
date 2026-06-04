#!/usr/bin/env python3
"""Datagate - Dataset conversion and query service."""

import argparse
import csv
import http.server
import json
import os
import re
import socket
import threading
import urllib.error
import urllib.parse
import urllib.request
from io import StringIO
from typing import Any

import xlrd
import openpyxl

# Configuration
STORAGE_DIR = "/tmp/datagate_storage"
DATASETS: dict[str, dict[str, Any]] = {}
STORAGE_LOCK = threading.Lock()


def load_datasets_from_storage() -> None:
    """Load persisted datasets from storage directory."""
    if not os.path.exists(STORAGE_DIR):
        return

    try:
        for filename in os.listdir(STORAGE_DIR):
            if filename.endswith('.json'):
                filepath = os.path.join(STORAGE_DIR, filename)
                try:
                    with open(filepath, 'r') as f:
                        data = json.load(f)
                    dataset_id = filename[:-5]  # Remove .json extension
                    # Validate the data structure
                    if 'columns' in data and 'rows' in data:
                        DATASETS[dataset_id] = data
                except (json.JSONDecodeError, KeyError):
                    # Skip invalid files
                    continue
    except OSError:
        pass


def save_datasets_to_storage() -> None:
    """Save all datasets to storage directory."""
    try:
        os.makedirs(STORAGE_DIR, exist_ok=True)
        for dataset_id, data in DATASETS.items():
            filepath = os.path.join(STORAGE_DIR, f"{dataset_id}.json")
            try:
                with open(filepath, 'w') as f:
                    json.dump(data, f)
            except (OSError, TypeError):
                # Skip datasets that can't be serialized
                continue
    except OSError:
        pass


def detect_file_type(content: bytes, filename: str = "") -> str:
    """Detect the file type based on content and extension."""
    filename_lower = filename.lower()

    # Check extension first
    if filename_lower.endswith(('.csv', '.tsv')):
        return 'csv'
    if filename_lower.endswith(('.xlsx', '.xls')):
        return 'excel'

    # Check content for Excel (ZIP-based format - starts with PK)
    if len(content) >= 4 and content[:2] == b'PK':
        return 'excel'

    # Default to CSV for text-based content
    return 'csv'


def infer_column_type(values: list[Any]) -> str:
    """Infer the type of a column based on its values."""
    if not values:
        return 'text'

    valid_count = 0
    int_count = 0
    float_count = 0

    for v in values:
        if v is None or v == '':
            continue
        try:
            # Try int first
            int(v)
            int_count += 1
            valid_count += 1
        except (ValueError, TypeError):
            try:
                # Try float
                float(v)
                float_count += 1
                valid_count += 1
            except (ValueError, TypeError):
                pass

    if valid_count == 0:
        return 'text'

    # If most valid values are integers, classify as integer
    if int_count > valid_count * 0.8:
        return 'integer'
    # If most valid values are floats, classify as float
    elif float_count > valid_count * 0.8:
        return 'float'
    else:
        return 'text'


def extract_csv_metadata(content: bytes) -> dict[str, Any]:
    """Extract metadata from CSV content."""
    # Decode content
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        # Try other common encodings
        for encoding in ['latin-1', 'iso-8859-1', 'cp1252']:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode('utf-8', errors='replace')

    # Parse CSV
    reader = csv.reader(StringIO(text))
    rows = list(reader)

    if not rows:
        return {
            'dataset_summary': {
                'filetype': 'csv',
                'row_count': 0,
                'column_count': 0
            },
            'column_details': {}
        }

    # First row is header
    headers = rows[0] if rows else []
    data_rows = rows[1:]

    # Build column details
    column_details = {}
    for idx, header in enumerate(headers):
        col_name = header if header else f'column_{idx}'
        values = []
        missing_count = 0

        for row in data_rows:
            if idx < len(row):
                val = row[idx].strip() if row[idx] else ''
                values.append(val if val else None)
                if not val:
                    missing_count += 1
            else:
                values.append(None)
                missing_count += 1

        distinct_values = set(v for v in values if v is not None)
        column_details[col_name] = {
            'type': infer_column_type([v for v in values if v is not None]),
            'distinct_count': len(distinct_values),
            'missing_count': missing_count
        }

    return {
        'dataset_summary': {
            'filetype': 'csv',
            'row_count': len(data_rows),
            'column_count': len(headers)
        },
        'column_details': column_details
    }


def extract_excel_metadata(content: bytes) -> dict[str, Any]:
    """Extract metadata from Excel content."""
    try:
        # Try openpyxl for xlsx
        workbook = openpyxl.loadWorkbook(StringIO(content.decode('latin-1')), data_only=True)
        if hasattr(workbook, 'worksheets') and workbook.worksheets:
            sheet = workbook.worksheets[0]
            max_row = sheet.max_row
            max_col = sheet.max_column
        else:
            max_row = 0
            max_col = 0
    except Exception:
        try:
            # Try xlrd for older Excel formats
            workbook = xlrd.open_workbook(file_contents=content)
            if workbook.sheets():
                sheet = workbook.sheet_by_index(0)
                max_row = sheet.nrows
                max_col = sheet.ncols
            else:
                max_row = 0
                max_col = 0
        except Exception:
            # Can't read Excel, provide basic info based on file size
            max_row = 0
            max_col = 0

    return {
        'dataset_summary': {
            'filetype': 'excel'
        }
    }


def ingest_dataset(content: bytes, url: str, enrich: bool = False) -> dict[str, Any] | tuple[int, str]:
    """Ingest a dataset from content and return structured data."""
    filename = urllib.parse.urlparse(url).path or "dataset"
    file_type = detect_file_type(content, filename)

    if file_type == 'csv':
        # Parse CSV
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            for encoding in ['latin-1', 'iso-8859-1', 'cp1252']:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                text = content.decode('utf-8', errors='replace')

        reader = csv.reader(StringIO(text))
        rows = list(reader)

        if not rows:
            return 400, "Empty CSV file"

        headers = rows[0] if rows else []
        data_rows = rows[1:]

        # Create dataset
        dataset_id = f"ds_{abs(hash(url))}_{len(DATASETS)}"
        dataset = {
            'columns': headers if headers else [f'column_{i}' for i in range(len(data_rows[0]) if data_rows else 0)],
            'rows': data_rows,
            'metadata': {
                'source_url': url,
                'file_type': 'csv'
            }
        }

        # Add enrichment if requested
        if enrich:
            metadata = extract_csv_metadata(content)
            dataset['metadata']['enrichment'] = metadata

        return 200, dataset_id, dataset

    elif file_type == 'excel':
        # Handle Excel files
        try:
            workbook = openpyxl.loadWorkbook(StringIO(content.decode('latin-1')), data_only=True)
            if hasattr(workbook, 'worksheets') and workbook.worksheets:
                sheet = workbook.worksheets[0]
                headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
                rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    rows.append([str(c) if c is not None else '' for c in row])
            else:
                headers = []
                rows = []
        except Exception:
            try:
                workbook = xlrd.open_workbook(file_contents=content)
                if workbook.sheets():
                    sheet = workbook.sheet_by_index(0)
                    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)] if sheet.nrows > 0 else []
                    rows = []
                    for row_idx in range(1, sheet.nrows):
                        row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                        rows.append([str(c) if c is not None else '' for c in row])
                else:
                    headers = []
                    rows = []
            except Exception:
                return 400, "Unsupported Excel file format"

        dataset_id = f"ds_{abs(hash(url))}_{len(DATASETS)}"
        dataset = {
            'columns': [str(h) if h else f'column_{i}' for i, h in enumerate(headers)] if headers else [],
            'rows': rows,
            'metadata': {
                'source_url': url,
                'file_type': 'excel'
            }
        }

        # Add enrichment if requested (Excel only gets dataset_summary)
        if enrich:
            metadata = extract_excel_metadata(content)
            dataset['metadata']['enrichment'] = metadata

        return 200, dataset_id, dataset

    else:
        return 400, "Unsupported file type"


class DatagateHandler(http.server.BaseHTTPRequestHandler):
    """HTTP request handler for Datagate."""

    def log_message(self, format: str, *args: Any) -> None:
        """Override to suppress default logging."""
        pass

    def send_json_response(self, status: int, data: dict[str, Any]) -> None:
        """Send a JSON response."""
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode('utf-8'))

    def do_OPTIONS(self) -> None:
        """Handle CORS preflight requests."""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', '*')
        self.end_headers()

    def do_GET(self) -> None:
        """Handle GET requests."""
        parsed_path = urllib.parse.urlparse(self.path)
        path = parsed_path.path
        query_params = urllib.parse.parse_qs(parsed_path.query)

        try:
            if path == '/convert':
                self.handle_convert(query_params)
            elif path.startswith('/datasets/'):
                dataset_id = path.split('/')[2]
                self.handle_dataset(dataset_id, query_params)
            else:
                self.send_json_response(404, {'error': 'Not found'})
        except Exception as e:
            self.send_json_response(500, {'error': str(e)})

    def handle_convert(self, query_params: dict[str, list[str]]) -> None:
        """Handle /convert endpoint."""
        # Check for source parameter
        if 'source' not in query_params:
            self.send_json_response(400, {'error': 'Missing source parameter'})
            return

        source_url = query_params['source'][0]

        # Validate URL
        parsed_url = urllib.parse.urlparse(source_url)
        if not parsed_url.scheme or not parsed_url.netloc:
            self.send_json_response(400, {'error': 'Invalid URL'})
            return

        # Check for charset parameter (optional, currently ignored)
        if 'charset' in query_params:
            charset = query_params['charset'][0]
            # Validate charset is a valid encoding name (basic check)
            if not re.match(r'^[a-zA-Z0-9\-]+$', charset):
                self.send_json_response(400, {'error': 'Invalid charset'})
                return

        # Check for enrich parameter
        enrich = query_params.get('enrich', ['no'])[0].lower() == 'yes'

        # Check for force parameter (forces re-ingestion)
        force = 'force' in query_params

        # Generate dataset ID for caching
        cache_key = f"{source_url}_{'enriched' if enrich else 'basic'}"
        dataset_id = f"ds_{abs(hash(cache_key))}"

        with STORAGE_LOCK:
            # Check if dataset exists and cached
            if not force and dataset_id in DATASETS:
                # Return cached dataset
                self.send_json_response(200, {
                    'ok': True,
                    'endpoint': f"/datasets/{dataset_id}"
                })
                return

            # Fetch content from URL
            try:
                req = urllib.request.Request(
                    source_url,
                    headers={'User-Agent': 'Datagate/1.0'}
                )
                with urllib.request.urlopen(req, timeout=30) as response:
                    content = response.read()
            except urllib.error.URLError as e:
                self.send_json_response(400, {'error': f'Failed to fetch URL: {str(e)}'})
                return
            except Exception as e:
                self.send_json_response(400, {'error': f'Error fetching URL: {str(e)}'})
                return

            # Ingest dataset
            result = ingest_dataset(content, source_url, enrich=enrich)

            if isinstance(result, tuple):
                status, error_msg = result
                self.send_json_response(status, {'error': error_msg})
                return

            status, returned_dataset_id, dataset = result

            # Store dataset
            DATASETS[returned_dataset_id] = dataset
            save_datasets_to_storage()

            # Return response
            self.send_json_response(200, {
                'ok': True,
                'endpoint': f"/datasets/{returned_dataset_id}"
            })

    def handle_dataset(self, dataset_id: str, query_params: dict[str, list[str]]) -> None:
        """Handle /datasets/<id> endpoint."""
        with STORAGE_LOCK:
            if dataset_id not in DATASETS:
                self.send_json_response(404, {'error': 'Dataset not found'})
                return

            dataset = DATASETS[dataset_id]

        # Build response
        response = {
            'ok': True,
            'columns': dataset['columns'],
            'rows': dataset['rows'],
            'total': len(dataset['rows'])
        }

        # Add enrichment metadata if present
        if 'enrichment' in dataset.get('metadata', {}):
            enrichment = dataset['metadata']['enrichment']
            response['dataset_summary'] = enrichment.get('dataset_summary', {})
            response['column_details'] = enrichment.get('column_details', {})

        self.send_json_response(200, response)


def run_server(address: str, port: int) -> None:
    """Run the HTTP server."""
    server_address = (address, port)
    httpd = http.server.HTTPServer(server_address, DatagateHandler)
    print(f"Datagate server running on http://{address}:{port}")
    httpd.serve_forever()


def main() -> None:
    """Main entry point."""
    parser = argparse.ArgumentParser(description='Datagate - Dataset conversion service')
    parser.add_argument('command', choices=['start'], help='Command to run')
    parser.add_argument('--port', type=int, default=18080, help='Port to listen on')
    parser.add_argument('--address', default='127.0.0.1', help='Address to bind to')

    args = parser.parse_args()

    # Load existing datasets from storage
    load_datasets_from_storage()

    if args.command == 'start':
        run_server(args.address, args.port)


if __name__ == '__main__':
    main()