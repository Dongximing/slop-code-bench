#!/usr/bin/env python3
"""Datagate - CSV data gateway service."""

import argparse
import csv
import hashlib
import io
import mimetypes
import os
import re
import sys
import time
from urllib.parse import urlparse

import chardet
import requests
from flask import Flask, jsonify, request, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)
datasets = {}
DELIMITER_CHARS = [',', ';', '\t']

# Configuration defaults
DEFAULTS = {
    'MAX_SOURCE_SIZE': None,
    'ORIGIN_ALLOWLIST': None,
    'REQUIRE_TLS': False,
    'STORAGE_DIR': None,
    'CACHE_ENABLED': True,
}

# Runtime configuration
config = DEFAULTS.copy()

# Comparator registry for filters - maps comparator name to (validation, apply) functions
_COMPARATORS = {}


def parse_boolean(value, var_name=None):
    """Parse a boolean value strictly: 1/true/yes/on or 0/false/no/off (case-insensitive)."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ('1', 'true', 'yes', 'on'):
        return True
    if s in ('0', 'false', 'no', 'off'):
        return False
    raise ValueError(f"Invalid boolean value for {var_name or 'variable'}: {value}")


def parse_integer(value, var_name=None):
    """Parse an integer value."""
    if value is None:
        return None
    try:
        return int(value)
    except ValueError:
        raise ValueError(f"Invalid integer value for {var_name or 'variable'}: {value}")


def parse_list(value, var_name=None):
    """Parse a comma-separated list value."""
    if value is None:
        return None
    items = [item.strip() for item in value.split(',') if item.strip()]
    return items if items else None


def load_config_file(config_path):
    """Load configuration from a KEY=VALUE file.

    Format: KEY=VALUE, commas for list values, blank/comment lines ignored.
    Invalid config or read failure returns None.
    """
    if not os.path.exists(config_path):
        return None

    try:
        with open(config_path, 'r') as f:
            content = f.read()
    except (IOError, OSError):
        return None

    config = {}
    for line_num, line in enumerate(content.splitlines(), 1):
        line = line.strip()
        # Skip blank lines and comments
        if not line or line.startswith('#'):
            continue
        # Parse KEY=VALUE
        if '=' not in line:
            raise ValueError(f"Invalid config line {line_num}: missing '='")
        key, _, value = line.partition('=')
        key = key.strip()
        value = value.strip()
        if not key:
            raise ValueError(f"Invalid config line {line_num}: empty key")
        config[key] = value

    return config


def load_config():
    """Load configuration with precedence: 1. defaults, 2. config file, 3. env vars."""
    cfg = DEFAULTS.copy()

    # 1. Built-in defaults (already copied)

    # 2. DATAGATE_CONFIG file
    config_path = os.environ.get('DATAGATE_CONFIG')
    if config_path:
        try:
            file_config = load_config_file(config_path)
            if file_config is not None:
                cfg.update(file_config)
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)

    # 3. Direct environment variables (highest precedence)
    env_config = {}

    # MAX_SOURCE_SIZE
    max_size = os.environ.get('MAX_SOURCE_SIZE')
    if max_size is not None:
        try:
            env_config['MAX_SOURCE_SIZE'] = parse_integer(max_size, 'MAX_SOURCE_SIZE')
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)
    elif 'MAX_SOURCE_SIZE' in os.environ:
        # Empty string means unset
        env_config['MAX_SOURCE_SIZE'] = None

    # ORIGIN_ALLOWLIST
    origin_allowlist = os.environ.get('ORIGIN_ALLOWLIST')
    if origin_allowlist is not None:
        env_config['ORIGIN_ALLOWLIST'] = parse_list(origin_allowlist, 'ORIGIN_ALLOWLIST')
    elif 'ORIGIN_ALLOWLIST' in os.environ:
        env_config['ORIGIN_ALLOWLIST'] = None

    # REQUIRE_TLS
    require_tls = os.environ.get('REQUIRE_TLS')
    if require_tls is not None:
        try:
            env_config['REQUIRE_TLS'] = parse_boolean(require_tls, 'REQUIRE_TLS')
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)
    elif 'REQUIRE_TLS' in os.environ:
        env_config['REQUIRE_TLS'] = None

    # STORAGE_DIR
    storage_dir = os.environ.get('STORAGE_DIR')
    if storage_dir is not None:
        env_config['STORAGE_DIR'] = storage_dir if storage_dir != '' else None
    elif 'STORAGE_DIR' in os.environ:
        env_config['STORAGE_DIR'] = None

    # CACHE_ENABLED
    cache_enabled = os.environ.get('CACHE_ENABLED')
    if cache_enabled is not None:
        try:
            env_config['CACHE_ENABLED'] = parse_boolean(cache_enabled, 'CACHE_ENABLED')
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)
    elif 'CACHE_ENABLED' in os.environ:
        env_config['CACHE_ENABLED'] = None

    cfg.update(env_config)

    # Post-processing: ensure proper types
    # MAX_SOURCE_SIZE - parse from string if needed (from config file)
    if cfg.get('MAX_SOURCE_SIZE') is not None and not isinstance(cfg['MAX_SOURCE_SIZE'], int):
        try:
            cfg['MAX_SOURCE_SIZE'] = parse_integer(cfg['MAX_SOURCE_SIZE'], 'MAX_SOURCE_SIZE')
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)

    # ORIGIN_ALLOWLIST - parse from string if needed (from config file)
    if cfg.get('ORIGIN_ALLOWLIST') is not None and not isinstance(cfg['ORIGIN_ALLOWLIST'], list):
        cfg['ORIGIN_ALLOWLIST'] = parse_list(cfg['ORIGIN_ALLOWLIST'], 'ORIGIN_ALLOWLIST')

    # REQUIRE_TLS - parse from string if needed (from config file)
    if cfg.get('REQUIRE_TLS') is not None and not isinstance(cfg['REQUIRE_TLS'], bool):
        try:
            cfg['REQUIRE_TLS'] = parse_boolean(cfg['REQUIRE_TLS'], 'REQUIRE_TLS')
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)

    # STORAGE_DIR - ensure path is a string
    if cfg.get('STORAGE_DIR') is not None and not isinstance(cfg['STORAGE_DIR'], str):
        cfg['STORAGE_DIR'] = str(cfg['STORAGE_DIR'])

    # CACHE_ENABLED - parse from string if needed (from config file)
    if cfg.get('CACHE_ENABLED') is not None and not isinstance(cfg['CACHE_ENABLED'], bool):
        try:
            cfg['CACHE_ENABLED'] = parse_boolean(cfg['CACHE_ENABLED'], 'CACHE_ENABLED')
        except ValueError as e:
            print(f"Configuration error: {e}", file=sys.stderr)
            sys.exit(1)

    return cfg


# Load configuration at startup
try:
    config = load_config()
except SystemExit:
    raise
except Exception as e:
    print(f"Configuration error: {e}", file=sys.stderr)
    sys.exit(1)

# Initialize storage directory if configured
if config.get('STORAGE_DIR'):
    storage_path = config['STORAGE_DIR']
    try:
        os.makedirs(storage_path, exist_ok=True)
    except OSError as e:
        print(f"Failed to create storage directory {storage_path}: {e}", file=sys.stderr)
        sys.exit(1)

# Extract for easier access
MAX_SOURCE_SIZE = config.get('MAX_SOURCE_SIZE')
ORIGIN_ALLOWLIST = config.get('ORIGIN_ALLOWLIST')
REQUIRE_TLS = config.get('REQUIRE_TLS', False)
STORAGE_DIR = config.get('STORAGE_DIR')
CACHE_ENABLED = config.get('CACHE_ENABLED', True)


def check_origin_allowed():
    """Check if the request's Referer is allowed based on ORIGIN_ALLOWLIST.

    If ORIGIN_ALLOWLIST is not set, all requests pass.
    Otherwise:
    1. Require Referer header
    2. Parse hostname
    3. Accept only hostnames matching any allowed suffix case-insensitively with domain-boundary rules
    4. Return error dict if not allowed, None if allowed
    """
    if ORIGIN_ALLOWLIST is None:
        return None

    referer = request.headers.get('Referer') or request.headers.get('referer')
    if not referer:
        return {
            "ok": False,
            "error": "Missing Referer header"
        }

    # Parse the hostname from the Referer URL
    try:
        parsed = urlparse(referer)
        hostname = parsed.hostname
        if not hostname:
            return {
                "ok": False,
                "error": "Referer hostname not parseable"
            }
    except Exception:
        return {
            "ok": False,
            "error": "Referer hostname not parseable"
        }

    # Check against each allowed suffix with domain-boundary rules
    hostname_lower = hostname.lower()

    for suffix in ORIGIN_ALLOWLIST:
        suffix_lower = suffix.lower()
        # Domain-boundary: must match exactly at the end, preceded by a dot or start of string
        if hostname_lower == suffix_lower:
            return None
        if hostname_lower.endswith('.' + suffix_lower):
            return None

    # No match found
    return {
        "ok": False,
        "error": "Referer not allowed"
    }


def require_origin_allowlist(f):
    """Decorator to enforce origin allowlist check on a route."""
    def decorated(*args, **kwargs):
        result = check_origin_allowed()
        if result is not None:
            return jsonify(result), 403
        return f(*args, **kwargs)
    decorated.__name__ = f.__name__
    return decorated


def enforce_max_source_size(content_length):
    """Enforce MAX_SOURCE_SIZE limit. Returns None if allowed, error dict if exceeded."""
    if MAX_SOURCE_SIZE is None:
        return None

    if content_length is not None and content_length > MAX_SOURCE_SIZE:
        return {
            "ok": False,
            "error": f"File size exceeds maximum allowed size of {MAX_SOURCE_SIZE} bytes"
        }

    return None


def get_endpoint_url(dataset_id):
    """Generate endpoint URL based on REQUIRE_TLS setting.

    REQUIRE_TLS=false: return relative endpoint
    REQUIRE_TLS=true: return absolute https:// URL using request host
    """
    if not REQUIRE_TLS:
        return f"/datasets/{dataset_id}"
    # Use https:// with request host
    host = request.host
    # Flask's request.host includes port if non-standard
    return f"https://{host}/datasets/{dataset_id}"


def comparator(name):
    """Register a comparator function."""
    def decorator(func):
        _COMPARATORS[name] = func
        return func
    return decorator


try:
    CACHE_ENABLED = parse_boolean(os.environ.get('CACHE_ENABLED', 'true'), 'CACHE_ENABLED')
except ValueError as e:
    print(f"Configuration error: {e}", file=sys.stderr)
    sys.exit(1)


def is_valid_url(url):
    try:
        result = urlparse(url)
        return bool(result.scheme in ('http', 'https') and result.netloc)
    except Exception:
        return False


def detect_delimiter(sample):
    try:
        return csv.Sniffer().sniff(sample, delimiters=DELIMITER_CHARS).delimiter
    except Exception:
        return ','


def infer_type(value):
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s or re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', s):
        return s
    try:
        return int(s)
    except ValueError:
        pass
    try:
        f = float(s)
        return f if f == f and f != float('inf') and f != float('-inf') else s
    except ValueError:
        return s


def detect_charset(content, provided_charset=None):
    if provided_charset:
        try:
            content.decode(provided_charset)
            return provided_charset
        except (LookupError, UnicodeDecodeError):
            raise ValueError(f"Unsupported or malformed charset: {provided_charset}")
    result = chardet.detect(content)
    return result.get('encoding', 'utf-8') or 'utf-8'


def fetch_csv(source_url, charset=None):
    try:
        response = requests.get(source_url, timeout=30)
        if response.status_code != 200:
            raise requests.RequestException(f"HTTP {response.status_code}")
    except requests.RequestException as e:
        raise ValueError(f"Source unreachable or remote HTTP error: {e}")
    content = response.content
    encoding = detect_charset(content, charset)
    return content.decode(encoding)


def parse_csv(text):
    if not text.strip():
        raise ValueError("Empty content")
    delimiter = detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if len(rows) < 2:
        raise ValueError("Non-tabular content or insufficient data (requires at least header and one data row)")
    columns = rows[0]
    data_rows = rows[1:]
    if not all(len(row) == len(columns) for row in data_rows):
        raise ValueError("Non-tabular content - inconsistent column counts")
    return columns, [[infer_type(cell) for cell in row] for row in data_rows]


def parse_excel(file_bytes, filename):
    try:
        if filename.endswith('.xls'):
            import xlrd
            file_stream = io.BytesIO(file_bytes)
            workbook = xlrd.open_workbook(file_buffer=file_stream)
        else:
            from openpyxl import load_workbook
            file_stream = io.BytesIO(file_bytes)
            workbook = load_workbook(file_stream, read_only=True, data_only=True)

        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError("First sheet must be tabular (header + at least one data row)")

        columns = [str(c) if c is not None else '' for c in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                converted = [infer_type(cell) for cell in row]
                data_rows.append(converted)

        if not data_rows:
            raise ValueError("First sheet must be tabular (header + at least one data row)")

        for row in data_rows:
            if len(row) != len(columns):
                raise ValueError("Non-tabular content - inconsistent column counts")

        return columns, data_rows

    except ImportError as e:
        raise ValueError(f"Missing library for Excel support: {e}")
    except Exception as e:
        if "Unsupported format" in str(e) or "Cannot understand" in str(e):
            raise ValueError("Unsupported or corrupted spreadsheet format")
        raise


# Comparator registrations
@comparator('exact')
def cmp_exact(a, b):
    return str(a) == b


@comparator('contains')
def cmp_contains(a, b):
    return b in str(a)


@comparator('less')
def cmp_less(a, b):
    try:
        return float(a) < float(b)
    except (ValueError, TypeError):
        return False


@comparator('greater')
def cmp_greater(a, b):
    try:
        return float(a) > float(b)
    except (ValueError, TypeError):
        return False


def validate_numeric(value, param_name):
    """Validate that a value can be converted to float."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def parse_sort_param():
    sort_col = request.args.get('_sort')
    sort_desc_col = request.args.get('_sort_desc')
    if sort_desc_col is not None:
        return sort_desc_col, True
    return sort_col if sort_col else None, False if sort_col else None


def check_repeated_params(keys):
    """Check for repeated control parameters, returning error or None."""
    for key in keys:
        if len(request.args.getlist(key)) > 1:
            return f"Repeated control parameter: {key}"
    return None


def parse_filter_params(columns):
    """Parse filter parameters from request.args. Returns (filters, error)."""
    valid_comparators = set(_COMPARATORS.keys())
    has_numeric_comparator = {'less', 'greater'}

    # Check for duplicate filter keys
    raw_keys = set()
    for pair in request.query_string.decode('utf-8').split('&'):
        if '=' in pair:
            key = pair.split('=')[0]
            if not key.startswith('_') and '__' in key and key in raw_keys:
                return None, f"Duplicate filter key: {key}"
            raw_keys.add(key)

    filters = []
    for key, value in request.args.items():
        if key.startswith('_') or '__' not in key:
            continue

        column, comparator = key.split('__', 1)
        if comparator not in valid_comparators:
            return None, f"Invalid comparator: {comparator}"
        if column not in columns:
            return None, f"Unknown filter column: {column}"
        if comparator in has_numeric_comparator and not validate_numeric(value, key):
            return None, f"Comparator target not numeric: {key}"

        filters.append((column, comparator, value))

    return filters, None


def apply_filters(rows, columns, filters):
    """Apply filter list to rows, returning filtered rows."""
    if not filters:
        return rows
    col_indices = {col: columns.index(col) for col in {f[0] for f in filters}}

    def row_passes(row):
        for column, comparator, value in filters:
            comparator_func = _COMPARATORS[comparator]
            if not comparator_func(row[col_indices[column]], value):
                return False
        return True

    return [row for row in rows if row_passes(row)]


def sort_rows(rows, sort_col, columns, sort_desc):
    """Sort rows by column, returning sorted rows."""
    if sort_col not in columns:
        raise ValueError(f"Unknown column: {sort_col}")
    return sorted(rows, key=lambda x: x[columns.index(sort_col)], reverse=sort_desc)


def format_rows(rows, columns, shape, show_rowid, offset=0):
    """Format rows according to shape and display options."""
    if shape == 'objects':
        result = []
        for i, row in enumerate(rows, start=offset + 1):
            row_obj = {}
            if show_rowid:
                row_obj['rowid'] = i
            for ci, col in enumerate(columns):
                row_obj[col] = row[ci]
            result.append(row_obj)
        return result
    return [list(row) for row in rows]


def validate_int_param(param, name, min_val=None, max_val=None):
    """Validate integer parameter. Returns (value, error) tuple."""
    if param is None:
        return None, None
    try:
        val = int(param)
        if min_val is not None and val < min_val:
            return None, f"Invalid {name}: {param}"
        if max_val is not None and val > max_val:
            return None, f"Invalid {name}: {param}"
        return val, None
    except ValueError:
        return None, f"Invalid value: {param}"


def process_dataset(columns, rows, shape='lists', show_rowid=True, show_total=True):
    """Core processing: filters, sorts, paginates, formats. Returns (response_dict, total, status_code)."""
    start = time.time()

    # Validate and set defaults
    size, error = validate_int_param(request.args.get('_size'), '_size', min_val=1)
    if error:
        return {"ok": False, "error": error}, None, 400
    size = size if size is not None else 100

    offset, error = validate_int_param(request.args.get('_offset'), '_offset', min_val=0)
    if error:
        return {"ok": False, "error": error}, None, 400
    offset = offset if offset is not None else 0

    # Validate shape
    shape = request.args.get('_shape', 'lists')
    if shape not in ('lists', 'objects'):
        return {"ok": False, "error": f"Invalid value: {shape}"}, None, 400

    # Check repeated params
    repeated = check_repeated_params(('_size', '_offset', '_sort', '_sort_desc', '_shape', '_rowid', '_total'))
    if repeated:
        return {"ok": False, "error": repeated}, None, 400

    # Parse and apply filters
    filters, error = parse_filter_params(columns)
    if error:
        return {"ok": False, "error": error}, None, 400

    filtered = apply_filters(rows, columns, filters)
    total = len(filtered)

    # Sort if requested
    sort_col, sort_desc = parse_sort_param()
    if sort_col is not None:
        if sort_col == '':
            return {"ok": False, "error": "Empty sort column"}, None, 400
        try:
            filtered = sort_rows(filtered, sort_col, columns, sort_desc)
        except ValueError as e:
            return {"ok": False, "error": str(e)}, None, 400

    # Paginate and format
    paginated = filtered[offset:offset + size]
    result_rows = format_rows(paginated, columns, shape, show_rowid, offset)

    response = {"ok": True, "columns": columns, "rows": result_rows}
    if show_total:
        response["total"] = total
    response["query_ms"] = (time.time() - start) * 1000
    return response, total, 200


@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response


def compute_file_hash(file_stream):
    hasher = hashlib.md5()
    while True:
        chunk = file_stream.read(8192)
        if not chunk:
            break
        hasher.update(chunk)
    return hasher.hexdigest()


@app.route('/upload', methods=['POST'])
@require_origin_allowlist
def upload():
    # Check size limit
    content_length = request.content_length
    error = enforce_max_source_size(content_length)
    if error is not None:
        return jsonify(error), 400

    if request.content_type and 'multipart/form-data' not in request.content_type:
        return jsonify({"ok": False, "error": "Unsupported Media Type"}), 415

    uploaded_file = request.files.get('file') or request.files.get('attachment')

    if not uploaded_file or uploaded_file.filename == '':
        return jsonify({"ok": False, "error": "Missing file field"}), 400

    # Check the actual file size from the stream
    file_stream = uploaded_file.stream
    current_pos = file_stream.tell()
    file_stream.seek(0, io.SEEK_END)
    actual_size = file_stream.tell()
    file_stream.seek(current_pos, io.SEEK_SET)

    error = enforce_max_source_size(actual_size)
    if error is not None:
        return jsonify(error), 400

    filename = secure_filename(uploaded_file.filename)
    file_ext = mimetypes.guess_extension(uploaded_file.mimetype) or ''

    if filename.endswith('.csv'):
        format_type = 'csv'
    elif filename.endswith('.xls') or filename.endswith('.xlsx') or file_ext in ['.xls', '.xlsx']:
        format_type = 'excel'
    elif file_ext == '.csv':
        format_type = 'csv'
    else:
        return jsonify({"ok": False, "error": "Unsupported format: expected CSV, XLS, or XLSX"}), 400

    file_stream = uploaded_file.stream
    file_hash = compute_file_hash(file_stream)

    if file_hash in datasets:
        dataset_id = file_hash
        if 'source' not in datasets[dataset_id]:
            datasets[dataset_id]['source'] = f'file_upload_{filename}'
        return jsonify({"ok": True, "endpoint": get_endpoint_url(dataset_id)})

    file_stream.seek(0)

    try:
        if format_type == 'csv':
            content_bytes = file_stream.read()
            charset = request.args.get('charset')
            if charset:
                try:
                    text = content_bytes.decode(charset)
                except (LookupError, UnicodeDecodeError):
                    raise ValueError(f"Unsupported or malformed charset: {charset}")
            else:
                detected = detect_charset(content_bytes)
                text = content_bytes.decode(detected)
            columns, rows = parse_csv(text)
        else:
            content_bytes = file_stream.read()
            columns, rows = parse_excel(content_bytes, filename)
    except ValueError as e:
        error_msg = str(e)
        if 'Unsupported or corrupted' in error_msg:
            return jsonify({"ok": False, "error": error_msg}), 400
        raise

    dataset_id = file_hash
    datasets[dataset_id] = {'columns': columns, 'rows': rows, 'source': f'file_upload_{filename}'}
    return jsonify({"ok": True, "endpoint": get_endpoint_url(dataset_id)})


@app.route('/convert', methods=['GET'])
@require_origin_allowlist
def convert():
    source = request.args.get('source')
    if not source:
        return jsonify({"ok": False, "error": "Missing source parameter"}), 400
    if not is_valid_url(source):
        return jsonify({"ok": False, "error": "Invalid URL"}), 400

    force_values = request.args.getlist('force')
    if len(force_values) > 1:
        return jsonify({"ok": False, "error": "Invalid force parameter"}), 400
    force_reingest = CACHE_ENABLED and bool(force_values)

    dataset_id = hashlib.md5(source.encode()).hexdigest()

    if CACHE_ENABLED and not force_reingest and dataset_id in datasets:
        return jsonify({"ok": True, "endpoint": get_endpoint_url(dataset_id)})

    try:
        response = requests.get(source, timeout=30)
        if response.status_code != 200:
            raise requests.RequestException(f"HTTP {response.status_code}")
        content = response.content

        # Check content size
        error = enforce_max_source_size(len(content))
        if error is not None:
            return jsonify(error), 400

        content_type = response.headers.get('Content-Type', '').lower()
        is_excel = any(ext in source.lower() for ext in ['.xls', '.xlsx']) or \
                   'excel' in content_type or \
                   'vnd.ms-excel' in content_type or \
                   'vnd.openxmlformats-officedocument.spreadsheet' in content_type

        if is_excel:
            columns, rows = parse_excel(content, 'source_file.xlsx')
        else:
            charset = request.args.get('charset')
            if charset:
                try:
                    text = content.decode(charset)
                except (LookupError, UnicodeDecodeError):
                    raise ValueError(f"Unsupported or malformed charset: {charset}")
            else:
                encoding = detect_charset(content)
                text = content.decode(encoding)
            columns, rows = parse_csv(text)

    except requests.RequestException as e:
        error_msg = str(e)
        if dataset_id in datasets:
            return jsonify({"ok": True, "endpoint": get_endpoint_url(dataset_id)})
        return jsonify({"ok": False, "error": f"Source unreachable or remote HTTP error: {error_msg}"}), 404
    except ValueError as e:
        error_msg = str(e)
        if dataset_id in datasets:
            return jsonify({"ok": True, "endpoint": get_endpoint_url(dataset_id)})
        return jsonify({"ok": False, "error": error_msg}), 400

    datasets[dataset_id] = {'columns': columns, 'rows': rows, 'source': source}
    return jsonify({"ok": True, "endpoint": get_endpoint_url(dataset_id)})


@app.route('/datasets/<dataset_id>/export', methods=['GET'])
def export_dataset(dataset_id):
    if dataset_id not in datasets:
        return jsonify({"ok": False, "error": f"Unknown dataset id: {dataset_id}"}), 404

    data = datasets[dataset_id]
    columns = data['columns']
    rows = data['rows']

    filter_error = check_repeated_params(('_sort', '_sort_desc'))
    if filter_error:
        return jsonify({"ok": False, "error": filter_error}), 400

    filters, error = parse_filter_params(columns)
    if error:
        return jsonify({"ok": False, "error": error}), 400

    filtered = apply_filters(rows, columns, filters)

    sort_col, sort_desc = parse_sort_param()
    if sort_col is not None:
        if sort_col == '':
            return jsonify({"ok": False, "error": "Empty sort column"}), 400
        try:
            filtered = sort_rows(filtered, sort_col, columns, sort_desc)
        except ValueError as e:
            return jsonify({"ok": False, "error": str(e)}), 400

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in filtered:
        writer.writerow(row)

    csv_bytes = output.getvalue().encode('utf-8')
    response = send_file(
        io.BytesIO(csv_bytes),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f"{dataset_id}.csv"
    )
    response.headers['Content-Disposition'] = f'attachment; filename="{dataset_id}.csv"'
    return response


@app.route('/datasets/<dataset_id>', methods=['GET'])
def get_dataset(dataset_id):
    if dataset_id not in datasets:
        return jsonify({"ok": False, "error": f"Unknown dataset id: {dataset_id}"}), 404

    data = datasets[dataset_id]
    columns = data['columns']
    rows = data['rows']

    shape_str = request.args.get('_shape', 'lists')

    rowid_str = request.args.get('_rowid')
    if rowid_str is not None and rowid_str not in ('hide',):
        return jsonify({"ok": False, "error": f"Invalid _rowid value: {rowid_str}"}), 400
    show_rowid = rowid_str != 'hide' if rowid_str is not None else True

    total_str = request.args.get('_total')
    if total_str is not None and total_str not in ('hide',):
        return jsonify({"ok": False, "error": f"Invalid _total value: {total_str}"}), 400
    show_total = total_str != 'hide' if total_str is not None else True

    response, _, status = process_dataset(columns, rows, shape=shape_str, show_rowid=show_rowid, show_total=show_total)
    return jsonify(response), status


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>', methods=['GET'])
def catch_all(path):
    return jsonify({"ok": False, "error": "Not found"}), 404


def main():
    parser = argparse.ArgumentParser(description='Datagate CSV Gateway')
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    start_parser = subparsers.add_parser('start', help='Start the server')
    start_parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    start_parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')
    args = parser.parse_args()
    if args.command == 'start':
        app.run(host=args.address, port=args.port, debug=False)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
