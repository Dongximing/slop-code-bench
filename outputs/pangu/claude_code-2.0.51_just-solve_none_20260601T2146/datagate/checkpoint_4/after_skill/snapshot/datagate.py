#!/usr/bin/env python3
"""Datagate - CSV data gateway service."""

import argparse
import csv
import hashlib
import io
import mimetypes
import re
import sys
import time
from urllib.parse import urlparse

import chardet
import requests
from flask import Flask, jsonify, request, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)
datasets = {}
DELIMITER_CHARS = [',', ';', '\t']


def is_valid_url(url):
    try:
        result = urlparse(url)
        return bool(result.scheme in ('http', 'https') and result.netloc)
    except Exception:
        return False


def detect_delimiter(sample):
    try:
        return csv.Sniffer().sniff(sample, delimiters=DELIMITER_CHARS).delimiter
    except Exception:
        return ','


def infer_type(value):
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s or re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', s):
        return s
    try:
        return int(s)
    except ValueError:
        pass
    try:
        f = float(s)
        return f if f == f and f != float('inf') and f != float('-inf') else s
    except ValueError:
        return s


def detect_charset(content, provided_charset=None):
    if provided_charset:
        try:
            content.decode(provided_charset)
            return provided_charset
        except (LookupError, UnicodeDecodeError):
            raise ValueError(f"Unsupported or malformed charset: {provided_charset}")
    result = chardet.detect(content)
    return result.get('encoding', 'utf-8') or 'utf-8'


def fetch_csv(source_url, charset=None):
    try:
        response = requests.get(source_url, timeout=30)
        if response.status_code != 200:
            raise requests.RequestException(f"HTTP {response.status_code}")
    except requests.RequestException as e:
        raise ValueError(f"Source unreachable or remote HTTP error: {e}")
    content = response.content
    encoding = detect_charset(content, charset)
    return content.decode(encoding)


def parse_csv(text):
    if not text.strip():
        raise ValueError("Empty content")
    delimiter = detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if len(rows) < 2:
        raise ValueError("Non-tabular content or insufficient data (requires at least header and one data row)")
    columns = rows[0]
    data_rows = rows[1:]
    if not all(len(row) == len(columns) for row in data_rows):
        raise ValueError("Non-tabular content - inconsistent column counts")
    return columns, [[infer_type(cell) for cell in row] for row in data_rows]


def parse_excel(file_bytes, filename):
    """Parse XLS or XLSX file, return (columns, rows)."""
    try:
        if filename.endswith('.xls'):
            import xlrd
            file_stream = io.BytesIO(file_bytes)
            workbook = xlrd.open_workbook(file_buffer=file_stream)
        else:  # .xlsx
            from openpyxl import load_workbook
            file_stream = io.BytesIO(file_bytes)
            workbook = load_workbook(file_stream, read_only=True, data_only=True)

        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError("First sheet must be tabular (header + at least one data row)")

        columns = [str(c) if c is not None else '' for c in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                converted = [infer_type(cell) for cell in row]
                data_rows.append(converted)

        if not data_rows:
            raise ValueError("First sheet must be tabular (header + at least one data row)")

        for row in data_rows:
            if len(row) != len(columns):
                raise ValueError("Non-tabular content - inconsistent column counts")

        return columns, data_rows

    except ImportError as e:
        raise ValueError(f"Missing library for Excel support: {e}")
    except Exception as e:
        if "Unsupported format" in str(e) or "Cannot understand" in str(e):
            raise ValueError("Unsupported or corrupted spreadsheet format")
        raise


def apply_filter(row, columns, filters):
    for column, comparator, value in filters:
        col_idx = columns.index(column)
        cell_value = row[col_idx]

        if comparator == 'exact':
            if str(cell_value) != value:
                return False
        elif comparator == 'contains':
            if value not in str(cell_value):
                return False
        elif comparator == 'less':
            try:
                cell_num = float(cell_value)
                if not cell_num < float(value):
                    return False
            except (ValueError, TypeError):
                return False
        elif comparator == 'greater':
            try:
                cell_num = float(cell_value)
                if not cell_num > float(value):
                    return False
            except (ValueError, TypeError):
                return False
    return True


def parse_sort_param():
    """Parse sort parameters. _sort_desc wins if both present. Returns (column, desc) or None."""
    sort_col = request.args.get('_sort')
    sort_desc_col = request.args.get('_sort_desc')
    if sort_desc_col is not None:
        return sort_desc_col, True
    if sort_col is not None and sort_col:
        return sort_col, False
    return None, None


def parse_filter_params(columns):
    """Parse filter parameters from request.args. Returns (filters, error)."""
    filters = []
    valid_comparators = {'exact', 'contains', 'less', 'greater'}

    # Check for duplicate filter keys in raw query string
    query = request.query_string.decode('utf-8')
    raw_keys = set()
    for pair in query.split('&'):
        if '=' in pair:
            key = pair.split('=')[0]
            if not key.startswith('_') and '__' in key:
                if key in raw_keys:
                    return None, f"Duplicate filter key: {key}"
                raw_keys.add(key)

    for key, value in request.args.items():
        if key.startswith('_') or '__' not in key:
            continue

        parts = key.split('__', 1)
        if len(parts) != 2:
            continue

        column, comparator = parts

        if comparator not in valid_comparators:
            return None, f"Invalid comparator: {comparator}"

        if column not in columns:
            return None, f"Unknown filter column: {column}"

        if comparator in ('less', 'greater'):
            try:
                float(value)
            except (ValueError, TypeError):
                return None, f"Comparator target not numeric: {key}"

        filters.append((column, comparator, value))

    return filters, None


def process_dataset(columns, rows, shape='lists', show_rowid=True, show_total=True):
    """Core processing: filters, sorts, paginates, formats. Returns (response_dict, total, status_code)."""
    start = time.time()

    # Parse and validate size/offset
    size_str = request.args.get('_size')
    offset_str = request.args.get('_offset')

    if size_str is not None:
        try:
            size = int(size_str)
            if size <= 0:
                return {"ok": False, "error": f"Invalid _size: {size_str}"}, None, 400
        except ValueError:
            return {"ok": False, "error": f"Invalid value: {size_str}"}, None, 400
    else:
        size = 100

    if offset_str is not None:
        try:
            offset = int(offset_str)
            if offset < 0:
                return {"ok": False, "error": f"Invalid _offset: {offset_str}"}, None, 400
        except ValueError:
            return {"ok": False, "error": f"Invalid value: {offset_str}"}, None, 400
    else:
        offset = 0

    # Check for repeated parameters
    for key in ('_size', '_offset', '_sort', '_sort_desc', '_shape', '_rowid', '_total'):
        values = request.args.getlist(key)
        if len(values) > 1:
            return {"ok": False, "error": f"Repeated control parameter: {key}"}, None, 400

    # Validate shape
    if shape not in ('lists', 'objects'):
        return {"ok": False, "error": f"Invalid value: {shape}"}, None, 400

    # Parse filters
    filters, error = parse_filter_params(columns)
    if error:
        return {"ok": False, "error": error}, None, 400

    # Apply filters
    filtered = [row for row in rows if apply_filter(row, columns, filters)]
    total = len(filtered)

    # Parse sort
    sort_column, sort_desc = parse_sort_param()
    if sort_column is not None:
        if sort_column == '':
            return {"ok": False, "error": "Empty sort column"}, None, 400
        if sort_column not in columns:
            return {"ok": False, "error": f"Unknown column: {sort_column}"}, None, 400
        col_idx = columns.index(sort_column)
        filtered = sorted(filtered, key=lambda x: x[col_idx], reverse=sort_desc)

    # Paginate
    rows_with_id = [(i + 1, row) for i, row in enumerate(filtered)]
    paginated = rows_with_id[offset:offset + size]

    # Format response
    if shape == 'objects':
        result_rows = []
        for rowid, row in paginated:
            row_obj = {}
            if show_rowid:
                row_obj['rowid'] = rowid
            for i, col in enumerate(columns):
                row_obj[col] = row[i]
            result_rows.append(row_obj)
    else:
        result_rows = [row for _, row in paginated]

    response = {"ok": True, "columns": columns, "rows": result_rows}
    if show_total:
        response["total"] = total
    response["query_ms"] = (time.time() - start) * 1000
    return response, total, 200


def get_dataset_data(dataset_id):
    """Helper to get dataset and apply filter/sort/pagination."""
    if dataset_id not in datasets:
        return None, None, None

    data = datasets[dataset_id]
    columns = data['columns']
    rows = data['rows']

    filters, error = parse_filter_params(columns)
    if error:
        return None, None, error

    filtered = [row for row in rows if apply_filter(row, columns, filters)]

    sort_column, sort_desc = parse_sort_param()
    if sort_column is not None:
        if sort_column == '':
            return None, None, "Empty sort column"
        if sort_column not in columns:
            return None, None, "Unknown column"
        col_idx = columns.index(sort_column)
        filtered = sorted(filtered, key=lambda x: x[col_idx], reverse=sort_desc)

    return columns, filtered, None


@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response


def compute_file_hash(file_stream):
    hasher = hashlib.md5()
    while True:
        chunk = file_stream.read(8192)
        if not chunk:
            break
        hasher.update(chunk)
    return hasher.hexdigest()


@app.route('/upload', methods=['POST'])
def upload():
    if not request.is_multipart:
        return jsonify({"ok": False, "error": "Unsupported Media Type"}), 415

    uploaded_file = request.files.get('file') or request.files.get('attachment')

    if not uploaded_file or uploaded_file.filename == '':
        return jsonify({"ok": False, "error": "Missing file field"}), 400

    filename = secure_filename(uploaded_file.filename)
    file_ext = mimetypes.guess_extension(uploaded_file.mimetype) or ''

    if filename.endswith('.csv'):
        format_type = 'csv'
    elif filename.endswith('.xls') or filename.endswith('.xlsx'):
        format_type = 'excel'
    elif file_ext in ['.xls', '.xlsx']:
        format_type = 'excel'
    elif file_ext == '.csv':
        format_type = 'csv'
    else:
        return jsonify({"ok": False, "error": "Unsupported format: expected CSV, XLS, or XLSX"}), 400

    file_stream = uploaded_file.stream
    file_hash = compute_file_hash(file_stream)

    if file_hash in datasets:
        dataset_id = file_hash
        if 'source' not in datasets[dataset_id]:
            datasets[dataset_id]['source'] = f'file_upload_{filename}'
        return jsonify({"ok": True, "endpoint": f"/datasets/{dataset_id}"})

    file_stream.seek(0)

    try:
        if format_type == 'csv':
            content_bytes = file_stream.read()
            charset = request.args.get('charset')
            if charset:
                try:
                    text = content_bytes.decode(charset)
                except (LookupError, UnicodeDecodeError):
                    raise ValueError(f"Unsupported or malformed charset: {charset}")
            else:
                detected = detect_charset(content_bytes)
                text = content_bytes.decode(detected)
            columns, rows = parse_csv(text)
        else:
            content_bytes = file_stream.read()
            columns, rows = parse_excel(content_bytes, filename)
    except ValueError as e:
        error_msg = str(e)
        if 'Unsupported or corrupted' in error_msg:
            return jsonify({"ok": False, "error": error_msg}), 400
        raise

    dataset_id = file_hash
    datasets[dataset_id] = {'columns': columns, 'rows': rows, 'source': f'file_upload_{filename}'}
    return jsonify({"ok": True, "endpoint": f"/datasets/{dataset_id}"})


@app.route('/convert', methods=['GET'])
def convert():
    source = request.args.get('source')
    if not source:
        return jsonify({"ok": False, "error": "Missing source parameter"}), 400
    if not is_valid_url(source):
        return jsonify({"ok": False, "error": "Invalid URL"}), 400

    try:
        response = requests.get(source, timeout=30)
        if response.status_code != 200:
            raise requests.RequestException(f"HTTP {response.status_code}")
        content = response.content

        content_type = response.headers.get('Content-Type', '').lower()
        is_excel = any(ext in source.lower() for ext in ['.xls', '.xlsx']) or \
                   'excel' in content_type or \
                   'vnd.ms-excel' in content_type or \
                   'vnd.openxmlformats-officedocument.spreadsheet' in content_type

        if is_excel:
            columns, rows = parse_excel(content, 'source_file.xlsx')
        else:
            charset = request.args.get('charset')
            if charset:
                try:
                    text = content.decode(charset)
                except (LookupError, UnicodeDecodeError):
                    raise ValueError(f"Unsupported or malformed charset: {charset}")
            else:
                encoding = detect_charset(content)
                text = content.decode(encoding)
            columns, rows = parse_csv(text)

    except requests.RequestException as e:
        error_msg = str(e)
        return jsonify({"ok": False, "error": f"Source unreachable or remote HTTP error: {error_msg}"}), 404
    except ValueError as e:
        error_msg = str(e)
        return jsonify({"ok": False, "error": error_msg}), 400

    dataset_id = hashlib.md5(source.encode()).hexdigest()
    datasets[dataset_id] = {'columns': columns, 'rows': rows, 'source': source}
    return jsonify({"ok": True, "endpoint": f"/datasets/{dataset_id}"})


@app.route('/datasets/<dataset_id>/export', methods=['GET'])
def export_dataset(dataset_id):
    if dataset_id not in datasets:
        return jsonify({"ok": False, "error": f"Unknown dataset id: {dataset_id}"}), 404

    columns, filtered_rows, error = get_dataset_data(dataset_id)
    if error:
        return jsonify({"ok": False, "error": error}), 400

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in filtered_rows:
        writer.writerow(row)

    csv_bytes = output.getvalue().encode('utf-8')
    response = send_file(
        io.BytesIO(csv_bytes),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f"{dataset_id}.csv"
    )
    response.headers['Content-Disposition'] = f'attachment; filename="{dataset_id}.csv"'
    return response


@app.route('/datasets/<dataset_id>', methods=['GET'])
def get_dataset(dataset_id):
    if dataset_id not in datasets:
        return jsonify({"ok": False, "error": f"Unknown dataset id: {dataset_id}"}), 404

    data = datasets[dataset_id]
    columns = data['columns']
    rows = data['rows']

    shape_str = request.args.get('_shape', 'lists')
    rowid_str = request.args.get('_rowid')
    total_str = request.args.get('_total')

    # Validate rowid
    if rowid_str is not None and rowid_str not in ('hide',):
        return jsonify({"ok": False, "error": f"Invalid _rowid value: {rowid_str}"}), 400
    show_rowid = rowid_str != 'hide' if rowid_str is not None else True

    # Validate total
    if total_str is not None and total_str not in ('hide',):
        return jsonify({"ok": False, "error": f"Invalid _total value: {total_str}"}), 400
    show_total = total_str != 'hide' if total_str is not None else True

    response, _, status = process_dataset(columns, rows, shape=shape_str, show_rowid=show_rowid, show_total=show_total)
    return jsonify(response), status


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>', methods=['GET'])
def catch_all(path):
    return jsonify({"ok": False, "error": "Not found"}), 404


def main():
    parser = argparse.ArgumentParser(description='Datagate CSV Gateway')
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    start_parser = subparsers.add_parser('start', help='Start the server')
    start_parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    start_parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')
    args = parser.parse_args()
    if args.command == 'start':
        app.run(host=args.address, port=args.port, debug=False)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
