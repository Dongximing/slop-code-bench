#!/usr/bin/env python3
"""datagate - A CSV conversion and dataset query service."""

import argparse
import csv
import hashlib
import io
import json
import os
import re
import sys
import time
from urllib.parse import urlparse

import chardet
import requests
from flask import Flask, Response, request

# For Excel support - optional dependency
try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# For xlrd support for .xls files
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# Pre-compiled regex for time pattern matching
_TIME_PATTERN = re.compile(r'^\d{1,2}:\d{2}$')


# Configuration defaults
DEFAULT_CONFIG = {
    'MAX_SOURCE_SIZE': None,
    'ORIGIN_ALLOWLIST': None,
    'REQUIRE_TLS': 'false',
    'STORAGE_DIR': './data',
    'CACHE_ENABLED': 'true',
}

# Configuration storage
_config = dict(DEFAULT_CONFIG)

# Parsed origin allowlist (list of lowercase domain suffixes)
_origin_allowlist = None

# Parse boolean values strictly: 1/true/yes/on -> true, 0/false/no/off -> false
def parse_bool(value):
    """Parse boolean value strictly.

    Accepts: 1/true/yes/on (case-insensitive) -> True
    Accepts: 0/false/no/off (case-insensitive) -> False
    Returns: bool or None if invalid
    """
    if value is None:
        return None
    normalized = value.strip().lower()
    true_values = {'1', 'true', 'yes', 'on'}
    false_values = {'0', 'false', 'no', 'off'}
    if normalized in true_values:
        return True
    elif normalized in false_values:
        return False
    return None


def parse_int(value):
    """Parse integer value. Returns None if invalid or not set."""
    if value is None or value.strip() == '':
        return None
    try:
        return int(value)
    except ValueError:
        return None


def parse_list(value):
    """Parse comma-separated list. Returns None if empty/unset."""
    if value is None or value.strip() == '':
        return None
    items = [item.strip() for item in value.split(',')]
    items = [item for item in items if item]  # Remove empty items
    return items if items else None


def load_config_file(filepath):
    """Load configuration from a file.

    Format: KEY=VALUE lines, commas for list values, blank/comment lines ignored.
    Returns dict of parsed config or raises ValueError on error.
    """
    if not os.path.exists(filepath):
        return {}

    config = {}
    with open(filepath, 'r') as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            # Skip blank lines and comments
            if not line or line.startswith('#'):
                continue
            # Parse KEY=VALUE
            if '=' not in line:
                raise ValueError(f"Invalid config line {line_num}: missing '='")
            key, _, value = line.partition('=')
            key = key.strip()
            value = value.strip()
            if not key:
                raise ValueError(f"Invalid config line {line_num}: empty key")
            config[key] = value
    return config


def initialize_config():
    """Initialize configuration from defaults, config file, and environment variables.

    Precedence: built-in defaults < config file < direct environment variables.
    """
    config = dict(DEFAULT_CONFIG)

    # 1. Load from config file if DATAGATE_CONFIG is set
    config_file = os.environ.get('DATAGATE_CONFIG')
    if config_file:
        try:
            file_config = load_config_file(config_file)
            config.update(file_config)
        except Exception as e:
            print(f"Error loading config file: {e}", file=sys.stderr)
            sys.exit(1)

    # 2. Override with environment variables (direct env vars have highest precedence)
    env_bool_settings = ['REQUIRE_TLS', 'CACHE_ENABLED']
    env_int_settings = ['MAX_SOURCE_SIZE']
    env_list_settings = ['ORIGIN_ALLOWLIST']
    env_path_settings = ['STORAGE_DIR']

    for key in env_bool_settings:
        val = os.environ.get(key)
        if val is not None:
            parsed = parse_bool(val)
            if parsed is None:
                print(f"Error: Invalid boolean value for {key}: {val}", file=sys.stderr)
                sys.exit(1)
            config[key] = 'true' if parsed else 'false'

    for key in env_int_settings:
        val = os.environ.get(key)
        if val is not None:
            parsed = parse_int(val)
            if parsed is None or parsed < 0:
                print(f"Error: Invalid integer value for {key}: {val}", file=sys.stderr)
                sys.exit(1)
            config[key] = str(parsed)

    for key in env_list_settings:
        val = os.environ.get(key)
        if val is not None:
            parsed = parse_list(val)
            if parsed is None:
                print(f"Error: Invalid list value for {key}: {val}", file=sys.stderr)
                sys.exit(1)
            config[key] = ','.join(parsed)

    for key in env_path_settings:
        val = os.environ.get(key)
        if val is not None:
            if not val.strip():
                print(f"Error: Invalid path value for {key}: empty", file=sys.stderr)
                sys.exit(1)
            config[key] = val.strip()

    return config


def get_config(key):
    """Get a configuration value by key."""
    return _config.get(key)


def get_config_bool(key):
    """Get a boolean configuration value."""
    return parse_bool(_config.get(key))


def get_config_int(key):
    """Get an integer configuration value."""
    return parse_int(_config.get(key))


def get_config_list(key):
    """Get a list configuration value."""
    return parse_list(_config.get(key))


# Initialize configuration at module load
_config = initialize_config()


def get_max_source_size():
    """Get MAX_SOURCE_SIZE as integer or None if unset."""
    return get_config_int('MAX_SOURCE_SIZE')


def get_origin_allowlist():
    """Get ORIGIN_ALLOWLIST as list of lowercase domain suffixes, or None."""
    global _origin_allowlist
    if _origin_allowlist is None:
        val = get_config_list('ORIGIN_ALLOWLIST')
        _origin_allowlist = val
    return _origin_allowlist


def get_require_tls():
    """Get REQUIRE_TLS as boolean."""
    return get_config_bool('REQUIRE_TLS') or False


def get_storage_dir():
    """Get STORAGE_DIR path."""
    return get_config('STORAGE_DIR')


def get_cache_enabled():
    """Get CACHE_ENABLED as boolean."""
    return get_config_bool('CACHE_ENABLED') or True


# Cache enabled flag - from configuration
_cache_enabled = get_cache_enabled()

# Storage for datasets: id -> data
datasets = {}

# Map source URL to dataset id for determinism
url_to_id = {}

# Map file content hash to dataset id for determinism (for upload endpoint)
file_to_id = {}


def get_storage_path():
    """Get the storage directory path."""
    return get_storage_dir()


def ensure_storage_dir():
    """Ensure the storage directory exists."""
    storage_dir = get_storage_path()
    if not os.path.exists(storage_dir):
        os.makedirs(storage_dir, exist_ok=True)


def save_dataset(dataset_id: str, dataset: dict):
    """Save a dataset to persistent storage."""
    ensure_storage_dir()
    storage_dir = get_storage_path()
    filepath = os.path.join(storage_dir, f"{dataset_id}.json")

    # Convert rows to lists (in case they're tuples)
    data = {
        'columns': list(dataset['columns']),
        'rows': [list(row) for row in dataset['rows']]
    }

    with open(filepath, 'w') as f:
        json.dump(data, f)


def load_dataset(dataset_id: str):
    """Load a dataset from persistent storage.

    Returns:
        dataset dict or None if not found
    """
    storage_dir = get_storage_path()
    filepath = os.path.join(storage_dir, f"{dataset_id}.json")

    if not os.path.exists(filepath):
        return None

    try:
        with open(filepath, 'r') as f:
            data = json.load(f)

        # Convert rows back to lists
        return {
            'columns': data['columns'],
            'rows': [list(row) for row in data['rows']]
        }
    except Exception:
        return None


def load_all_datasets():
    """Load all persisted datasets at startup.

    Returns:
        tuple: (datasets dict, url_to_id dict, file_to_id dict)
    """
    storage_dir = get_storage_path()

    if not os.path.exists(storage_dir):
        return {}, {}, {}

    datasets = {}
    url_to_id = {}
    file_to_id = {}

    for filename in os.listdir(storage_dir):
        if not filename.endswith('.json'):
            continue

        dataset_id = filename[:-5]  # Remove .json extension
        dataset = load_dataset(dataset_id)

        if dataset is not None:
            datasets[dataset_id] = dataset

            # Try to load metadata
            metafilepath = os.path.join(storage_dir, f"{dataset_id}.meta.json")
            if os.path.exists(metafilepath):
                try:
                    with open(metafilepath, 'r') as f:
                        meta = json.load(f)
                        if 'source_url' in meta:
                            url_to_id[meta['source_url']] = dataset_id
                        if 'file_hash' in meta:
                            file_to_id[meta['file_hash']] = dataset_id
                except Exception:
                    pass

    return datasets, url_to_id, file_to_id


def save_dataset_metadata(dataset_id: str, source_url: str = None, file_hash: str = None):
    """Save metadata for a dataset (source URL or file hash mapping)."""
    ensure_storage_dir()
    storage_dir = get_storage_path()

    meta = {}
    if source_url:
        meta['source_url'] = source_url
    if file_hash:
        meta['file_hash'] = file_hash

    if not meta:
        return

    metafilepath = os.path.join(storage_dir, f"{dataset_id}.meta.json")
    with open(metafilepath, 'w') as f:
        json.dump(meta, f)


def is_origin_allowed(referer: str) -> bool:
    """Check if the Referer hostname matches any allowed suffix.

    Rules:
    - Case-insensitive matching
    - Domain-boundary: example.com matches www.example.com but notexample.com
    """
    allowlist = get_origin_allowlist()
    if allowlist is None:
        return True  # No allowlist means all pass

    if not referer:
        return False

    try:
        parsed = urlparse(referer)
        hostname = parsed.hostname
        if not hostname:
            return False

        hostname_lower = hostname.lower()

        for allowed_suffix in allowlist:
            allowed_lower = allowed_suffix.lower()
            # Check domain boundary: either exact match or ends with .suffix
            if hostname_lower == allowed_lower:
                return True
            if hostname_lower.endswith('.' + allowed_lower):
                return True

        return False
    except Exception:
        return False


@app.before_request
def check_origin_allowlist():
    """Check Referer header against ORIGIN_ALLOWLIST if configured."""
    allowlist = get_origin_allowlist()
    if allowlist is None:
        return  # No allowlist configured, skip check

    # Get Referer header
    referer = request.headers.get('Referer') or request.headers.get('referer')

    if not referer:
        return json_error(403, "Missing Referer")

    if not is_origin_allowed(referer):
        return json_error(403, "Referer not allowed")


def generate_id(source_url: str) -> str:
    return hashlib.md5(source_url.encode('utf-8')).hexdigest()[:12]


def generate_file_id(file_bytes: bytes) -> str:
    return hashlib.md5(file_bytes).hexdigest()[:12]


def infer_delimiter(sample: str) -> str:
    """Infer CSV delimiter from sample content."""
    delimiters = [',', '\t', ';']
    sniffer = csv.Sniffer()

    # Try sniffer first
    try:
        dialect = sniffer.sniff(sample)
        return dialect.delimiter
    except csv.Error:
        pass

    # Manual detection
    best_delim = ','
    best_count = 0

    for delim in delimiters:
        count = sample.count(delim)
        if count > best_count:
            best_count = count
            best_delim = delim

    return best_delim


def is_time_like(value: str) -> bool:
    """Check if a string value looks like a time (e.g., 08:30, 9:15, 12:00)."""
    return bool(_TIME_PATTERN.match(value.strip()))


def convert_value(value: str):
    value = value.strip()

    if is_time_like(value):
        return value

    # Try integer
    try:
        return int(value)
    except ValueError:
        pass

    # Try float/decimal
    try:
        return float(value)
    except ValueError:
        pass

    # Return as string
    return value


def parse_csv(content: str, delimiter: str) -> dict:
    """Parse CSV content and return structured data."""
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)

    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("CSV must have at least one header row and one data row")

    columns = rows[0]
    data_rows = rows[1:]

    # Convert each row with type inference
    converted_rows = []
    for row in data_rows:
        converted_row = [convert_value(cell) for cell in row]
        converted_rows.append(converted_row)

    return {
        'columns': columns,
        'rows': converted_rows
    }


def fetch_csv(source_url: str, charset: str = None) -> tuple:
    """Fetch CSV from remote URL and return (raw_content, encoding)."""
    try:
        response = requests.get(source_url, timeout=30, stream=True)
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}")

        # Check size limit
        max_size = get_max_source_size()
        if max_size is not None:
            content_bytes = b''
            total_size = 0
            chunk_size = 8192
            for chunk in response.iter_content(chunk_size=chunk_size):
                total_size += len(chunk)
                if total_size > max_size:
                    raise ValueError(f"Source size exceeds maximum allowed size of {max_size} bytes")
                content_bytes += chunk
        else:
            content_bytes = response.content

    except Exception as e:
        raise

    if not content_bytes.strip():
        raise ValueError("Empty content")

    # Determine encoding
    if charset:
        try:
            content = content_bytes.decode(charset)
            encoding = charset
        except (UnicodeDecodeError, LookupError):
            raise ValueError(f"Invalid charset: {charset}")
    else:
        detected = chardet.detect(content_bytes)
        encoding = detected.get('encoding', 'utf-8')
        try:
            content = content_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            # Fallback to UTF-8
            encoding = 'utf-8'
            content = content_bytes.decode('utf-8', errors='replace')

    return content, encoding


def detect_file_type(file_bytes: bytes) -> str:
    """Detect file type from magic bytes.

    Returns: 'csv', 'xls', 'xlsx', or None if unknown.
    """
    if not file_bytes:
        return None

    # Check for ZIP (xlsx files are ZIP archives)
    if file_bytes[:2] == b'PK':
        # xlsx files start with PK ( ZIP format)
        return 'xlsx'

    # Check for .xls (OLE2 Compound Document)
    if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return 'xls'

    # For CSV, try to decode as text
    try:
        # Try UTF-8 with BOM
        if file_bytes[:3] == b'\xef\xbb\xbf':
            text = file_bytes[3:].decode('utf-8')
        else:
            text = file_bytes.decode('utf-8')

        # Check if it looks like a CSV (at least one comma or tab/semicolon)
        lines = text.strip().split('\n')
        if len(lines) >= 2:
            header = lines[0]
            if ',' in header or '\t' in header or ';' in header:
                # Verify data rows have consistent structure
                delimiters = [',', '\t', ';']
                for delim in delimiters:
                    if delim in header:
                        fields_in_header = header.split(delim)
                        # Check first few data rows
                        for line in lines[1:4]:
                            if line.strip():
                                fields = line.split(delim)
                                if len(fields) != len(fields_in_header):
                                    break
                        else:
                            return 'csv'
                        break
    except (UnicodeDecodeError, UnicodeError):
        pass

    return None


def parse_excel_xlsx(file_bytes: bytes) -> dict:
    """Parse .xlsx Excel file and return structured data."""
    if not HAS_OPENPYXL:
        raise ValueError("openpyxl not installed")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Get all rows from the first worksheet
    rows_data = []
    for row in ws.iter_rows(values_only=True):
        rows_data.append(list(row))

    if len(rows_data) < 2:
        raise ValueError("Excel must have at least one header row and one data row")

    columns = [str(c) if c is not None else "" for c in rows_data[0]]
    data_rows = rows_data[1:]

    # Convert each row with type inference
    converted_rows = []
    for row in data_rows:
        converted_row = []
        for cell in row:
            if cell is None:
                converted_row.append("")
            else:
                # Convert value with type inference
                converted_row.append(convert_value(str(cell)))
        converted_rows.append(converted_row)

    return {
        'columns': columns,
        'rows': converted_rows
    }


def parse_excel_xls(file_bytes: bytes) -> dict:
    """Parse .xls Excel file and return structured data."""
    if not HAS_XLRD:
        raise ValueError("xlrd not installed")

    book = xlrd.open_workbook(file_contents=file_bytes)
    ws = book.sheet_by_index(0)

    if ws.nrows < 2:
        raise ValueError("Excel must have at least one header row and one data row")

    # Get header row
    columns = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]
    data_rows = []

    # Get data rows
    for row_idx in range(1, ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell_value(row_idx, col_idx)
            row.append(convert_value(str(cell)))
        data_rows.append(row)

    return {
        'columns': columns,
        'rows': data_rows
    }


def make_endpoint_url(dataset_id: str) -> str:
    """Create the endpoint URL based on REQUIRE_TLS setting.

    - REQUIRE_TLS=false: return relative endpoint
    - REQUIRE_TLS=true: return absolute https:// URL using request host
    """
    if not get_require_tls():
        return f"/datasets/{dataset_id}"

    # Build absolute https URL with request host
    host = request.host

    return f"https://{host}/datasets/{dataset_id}"


def is_tabular(content: str) -> bool:
    """Check if content appears to be tabular data."""
    lines = content.strip().split('\n')
    if len(lines) < 2:
        return False

    header = lines[0]
    delimiters = [',', '\t', ';']

    for delim in delimiters:
        if delim in header:
            # Verify data rows have similar structure
            fields_in_header = header.split(delim)
            for line in lines[1:4]:  # Check first few data rows
                if line.strip():
                    fields = line.split(delim)
                    if len(fields) != len(fields_in_header):
                        return False
            return True

    return False


def convert_source(source, charset, force_present=None):
    """Internal function to convert a source URL into a dataset.

    Args:
        source: The URL to fetch and parse
        charset: Optional character encoding for CSV
        force_present: Optional pre-computed force flag (if None, will be parsed from request)

    Returns:
        Response tuple (status, body) for the conversion result
    """
    # Handle force parameter if not provided
    if force_present is None:
        force = request.args.get('force')
        if force is not None and force.strip() != '':
            return json_error(400, "Invalid force parameter")
        force_present = force is not None

    # Store previous dataset info for error handling
    previous_dataset = None
    previous_id = None
    if source in url_to_id:
        previous_id = url_to_id[source]
        previous_dataset = datasets.get(previous_id)

    # Determine if we should use cache
    use_cache = _cache_enabled and not force_present

    # Return cached result if available and caching enabled
    if use_cache and source in url_to_id:
        dataset_id = url_to_id[source]
        return json_success(200, {"ok": True, "endpoint": make_endpoint_url(dataset_id)})

    # Fetch the file - try as binary first to detect format
    fetch_start_time = time.time()
    try:
        response = requests.get(source, timeout=30, stream=True)
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}")

        # Check size limit
        max_size = get_max_source_size()
        if max_size is not None:
            file_bytes = b''
            total_size = 0
            chunk_size = 8192
            for chunk in response.iter_content(chunk_size=chunk_size):
                total_size += len(chunk)
                if total_size > max_size:
                    return json_error(400, f"Source size exceeds maximum allowed size of {max_size} bytes")
                file_bytes += chunk
        else:
            file_bytes = response.content
    except Exception as e:
        error_msg = str(e)
        # Remote HTTP 404 should return 404
        if "HTTP 404" in error_msg:
            return json_error(404, "Source unreachable")
        # Other errors are bad requests
        return json_error(400, error_msg)

    if not file_bytes.strip():
        return json_error(400, "Empty content")

    # Detect file type from magic bytes
    file_type = detect_file_type(file_bytes)

    if file_type is None:
        return json_error(400, "Unsupported format")

    # For CSV, charset applies
    if file_type == 'csv':
        # Determine encoding
        if charset:
            try:
                content = file_bytes.decode(charset)
                encoding = charset
            except (UnicodeDecodeError, LookupError):
                return json_error(400, f"Invalid charset: {charset}")
        else:
            detected = chardet.detect(file_bytes)
            encoding = detected.get('encoding', 'utf-8')
            try:
                content = file_bytes.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                encoding = 'utf-8'
                content = file_bytes.decode('utf-8', errors='replace')

        if not is_tabular(content):
            return json_error(400, "Non-tabular content")

        sample = content[:4096]
        delimiter = infer_delimiter(sample)

        try:
            dataset = parse_csv(content, delimiter)
        except ValueError as e:
            return json_error(400, str(e))

    elif file_type == 'xlsx':
        if not HAS_OPENPYXL:
            return json_error(400, "Unsupported format: .xlsx requires openpyxl")
        try:
            dataset = parse_excel_xlsx(file_bytes)
        except ValueError as e:
            return json_error(400, str(e))
        except Exception as e:
            return json_error(400, f"Invalid file: {str(e)}")

    elif file_type == 'xls':
        if not HAS_XLRD:
            return json_error(400, "Unsupported format: .xls requires xlrd")
        try:
            dataset = parse_excel_xls(file_bytes)
        except ValueError as e:
            return json_error(400, str(e))
        except Exception as e:
            return json_error(400, f"Invalid file: {str(e)}")

    else:
        return json_error(400, "Unsupported format")

    # Generate deterministic ID
    dataset_id = generate_id(source)

    # Store dataset (in-memory and persisted)
    datasets[dataset_id] = dataset
    url_to_id[source] = dataset_id
    save_dataset(dataset_id, dataset)
    save_dataset_metadata(dataset_id, source_url=source)

    return json_success(200, {"ok": True, "endpoint": make_endpoint_url(dataset_id)})


@app.route('/convert')
def convert():
    """Convert remote file to dataset.

    Supports CSV, .xls, and .xlsx from URLs.
    charset applies only to text CSV sources.
    """
    return convert_source(request.args.get('source'),
                          request.args.get('charset'),
                          None)


@app.route('/upload', methods=['POST'])
def upload():
    """Upload a file for conversion.

    Accepts multipart form with field 'file' or 'attachment'.
    Supports CSV, .xls, and .xlsx formats.
    """
    # Check for file field (in Flask, request.files exists even for non-multipart)
    # But we need to verify it's actually a multipart request
    # In Flask, request.files is only populated for multipart/form-data
    # If we get here but there's no files dict entry, it wasn't multipart
    # We check if we can access a file from the form

    # Check if files dict has any entries (multipart)
    # or try to get a file using .get() which returns None if not multipart
    try:
        uploaded_file = request.files.get('file') or request.files.get('attachment')
        if uploaded_file is None:
            # Check if files dict exists and is empty vs request isn't multipart
            # In Flask, request.files always exists for any request, but is empty for non-multipart
            # Also check content-type for safety
            content_type = request.content_type or ''
            if 'multipart' not in content_type.lower():
                return json_error(415, "Not a multipart request")
    except RuntimeError:
        return json_error(415, "Not a multipart request")

    if not uploaded_file:
        return json_error(400, "Missing 'file' field")

    # Read file content and check size limit
    max_size = get_max_source_size()
    if max_size is not None:
        # Read in chunks to check size without loading everything
        file_bytes = b''
        chunk_size = 8192
        total_size = 0
        while True:
            chunk = uploaded_file.read(chunk_size)
            if not chunk:
                break
            total_size += len(chunk)
            if total_size > max_size:
                return json_error(400, f"File size exceeds maximum allowed size of {max_size} bytes")
            file_bytes += chunk
    else:
        file_bytes = uploaded_file.read()

    if not file_bytes:
        return json_error(400, "Empty file")

    # Detect file type
    file_type = detect_file_type(file_bytes)

    if file_type is None:
        return json_error(400, "Unsupported format")

    # Check for unsupported formats
    if file_type == 'xls' and not HAS_XLRD:
        return json_error(400, "Unsupported format: .xls requires xlrd")
    if file_type == 'xlsx' and not HAS_OPENPYXL:
        return json_error(400, "Unsupported format: .xlsx requires openpyxl")

    # For CSV, we also accept without openpyxl or xlrd
    if file_type not in ('csv', 'xls', 'xlsx'):
        return json_error(400, "Unsupported format")

    # Generate deterministic ID based on file content
    dataset_id = generate_file_id(file_bytes)

    # Check if we already have this dataset
    if dataset_id in datasets:
        return json_success(200, {"ok": True, "endpoint": make_endpoint_url(dataset_id)})

    # Parse the file
    try:
        if file_type == 'csv':
            # Try to decode as UTF-8
            try:
                text = file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                # Try chardet for encoding detection
                detected = chardet.detect(file_bytes)
                encoding = detected.get('encoding', 'utf-8')
                text = file_bytes.decode(encoding, errors='replace')

            if not is_tabular(text):
                return json_error(400, "Non-tabular content")

            sample = text[:4096]
            delimiter = infer_delimiter(sample)
            dataset = parse_csv(text, delimiter)
        elif file_type == 'xlsx':
            dataset = parse_excel_xlsx(file_bytes)
        elif file_type == 'xls':
            dataset = parse_excel_xls(file_bytes)
        else:
            return json_error(400, "Unsupported format")
    except ValueError as e:
        return json_error(400, str(e))
    except Exception as e:
        return json_error(400, f"Invalid file: {str(e)}")

    # Store dataset (in-memory and persisted)
    datasets[dataset_id] = dataset
    file_to_id[dataset_id] = dataset_id
    save_dataset(dataset_id, dataset)
    save_dataset_metadata(dataset_id, file_hash=dataset_id)

    return json_success(200, {"ok": True, "endpoint": make_endpoint_url(dataset_id)})


@app.route('/convert', methods=['POST'])
def convert_post():
    """Convert remote CSV to dataset (POST version).

    Accepts multipart form with field 'url' or 'source' for remote URL,
    or field 'file' or 'attachment' for direct file upload.
    """
    # Check if there's a file upload
    if request.is_multipart:
        uploaded_file = request.files.get('file') or request.files.get('attachment')
        if uploaded_file:
            # Use the upload endpoint logic but return from here
            return upload()

    # Check for URL parameter (can be form field or query string)
    source = request.form.get('url') or request.form.get('source') or request.args.get('source')
    charset = request.form.get('charset') or request.args.get('charset')
    force = request.form.get('force') or request.args.get('force')

    if not source:
        return json_error(400, "Missing 'source' parameter")

    # Validate URL
    try:
        parsed = urlparse(source)
        if not parsed.scheme or not parsed.netloc:
            raise ValueError("Invalid URL")
    except Exception:
        return json_error(400, "Invalid URL")

    if charset is not None and charset.strip() == '':
        return json_error(400, "Invalid charset: empty string")

    # Validate force parameter - if present, must be absent or empty (presence flag)
    # Any additional force value is HTTP 400
    if force is not None and force.strip() != '':
        return json_error(400, "Invalid force parameter")

    force_present = force is not None

    # Check if we already have this dataset
    # When caching is disabled, or force is present, we proceed to re-download
    if source in url_to_id and not force_present and _cache_enabled:
        dataset_id = url_to_id[source]
        return json_success(200, {"ok": True, "endpoint": make_endpoint_url(dataset_id)})

    # Fetch the CSV
    try:
        content, encoding = fetch_csv(source, charset)
    except Exception as e:
        error_msg = str(e)
        # Remote HTTP 404 should return 404
        if "HTTP 404" in error_msg:
            return json_error(404, "Source unreachable")
        # Other errors are bad requests
        return json_error(400, error_msg)

    # Validate it's tabular
    if not is_tabular(content):
        return json_error(400, "Non-tabular content")

    # Infer delimiter and parse
    sample = content[:4096]  # Sample for delimiter detection
    delimiter = infer_delimiter(sample)

    try:
        dataset = parse_csv(content, delimiter)
    except ValueError as e:
        return json_error(400, str(e))

    # Generate deterministic ID
    dataset_id = generate_id(source)

    # Store dataset (in-memory and persisted)
    datasets[dataset_id] = dataset
    url_to_id[source] = dataset_id
    save_dataset(dataset_id, dataset)
    save_dataset_metadata(dataset_id, source_url=source)

    return json_success(200, {"ok": True, "endpoint": make_endpoint_url(dataset_id)})


@app.route('/datasets/<dataset_id>')
def get_dataset(dataset_id):
    """Retrieve dataset by ID with pagination, sorting, and response controls."""
    if dataset_id not in datasets:
        return json_error(404, "Dataset not found")

    start_time = time.time()

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = [list(row) for row in dataset['rows']]  # Create mutable copy

    # Parse control parameters
    args = request.args

    # Check for duplicate parameters (control and filter) using getlist
    for key in args.keys():
        values = request.args.getlist(key)
        if len(values) > 1:
            if key.startswith('_'):
                return json_error(400, f"Repeated parameter: {key}")
            else:
                # It's a filter key
                return json_error(400, f"Duplicate filter key: {key}")

    # --- Filter parsing and validation ---
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filter_params = {}
    numeric_filter_values = {}  # Store parsed numeric values for less/greater

    for key in args.keys():
        # Skip control params (start with _)
        if key.startswith('_'):
            continue
        # Check for column__comparator pattern
        if '__' in key:
            parts = key.split('__', 1)
            if len(parts) == 2:
                col_name, comparator = parts
                if comparator not in valid_comparators:
                    return json_error(400, f"Invalid comparator: {comparator}")
                if col_name not in columns:
                    return json_error(400, f"Unknown filter column: {col_name}")

                filter_value = request.args.getlist(key)[0]
                # Validate numeric filter values for less/greater
                if comparator in ('less', 'greater'):
                    try:
                        numeric_filter_values[key] = float(filter_value)
                    except (ValueError, TypeError):
                        return json_error(400, f"Invalid numeric value: {filter_value}")

                filter_params[key] = filter_value

    # --- Apply filters ---
    def apply_filters(row):
        for filter_key, filter_value in filter_params.items():
            col_name, comparator = filter_key.split('__', 1)
            col_idx = columns.index(col_name)
            cell_value = row[col_idx]

            if comparator == 'exact':
                if str(cell_value) != filter_value:
                    return False
            elif comparator == 'contains':
                if filter_value not in str(cell_value):
                    return False
            elif comparator == 'less' or comparator == 'greater':
                # Skip rows with non-numeric stored values
                try:
                    cell_num = float(cell_value)
                except (ValueError, TypeError):
                    return False
                filter_num = numeric_filter_values[filter_key]
                if comparator == 'less':
                    if not (cell_num < filter_num):
                        return False
                elif comparator == 'greater':
                    if not (cell_num > filter_num):
                        return False
        return True

    # Apply filters to rows
    filtered_rows = []
    for row in rows:
        if apply_filters(row):
            filtered_rows.append(row)

    rows = filtered_rows

    # Validate and get _size
    size = 100
    if '_size' in args:
        size_str = args['_size']
        if not size_str.isdigit() or int(size_str) <= 0:
            return json_error(400, f"Invalid '_size': {size_str}")
        size = int(size_str)

    # Validate and get _offset
    offset = 0
    if '_offset' in args:
        offset_str = args['_offset']
        if not offset_str.isdigit() or int(offset_str) < 0:
            return json_error(400, f"Invalid '_offset': {offset_str}")
        offset = int(offset_str)

    # Validate and get _shape
    shape = 'lists'
    if '_shape' in args:
        shape = args['_shape']
        if shape not in ['lists', 'objects']:
            return json_error(400, f"Invalid '_shape': {shape}")

    # Validate _sort/_sort_desc columns and get sort parameters
    sort_col = args.get('_sort_desc') or args.get('_sort')
    sort_desc = '_sort_desc' in args

    # Validate sort column
    if sort_col is not None:
        if sort_col == '' or sort_col not in columns:
            return json_error(400, f"Unknown column for sorting: {sort_col}")
        # Precompute column index mapping for O(1) lookups
        col_to_idx = {name: i for i, name in enumerate(columns)}
        col_idx = col_to_idx[sort_col]
        # Stable sort: for each row, use the sort column value, and for tie-breaking,
        # use the values of all columns in order
        def sort_key(row):
            key = [row[col_idx]]
            for c in columns:
                if c != sort_col:
                    key.append(row[col_to_idx[c]])
            return tuple(key)
        rows.sort(key=sort_key, reverse=sort_desc)

    total = len(rows)

    # Apply pagination
    paginated_rows = rows[offset:offset + size]

    # Format rows based on shape
    if shape == 'objects':
        formatted_rows = []
        for i, row in enumerate(paginated_rows):
            obj = {}
            for col_idx, col_name in enumerate(columns):
                obj[col_name] = row[col_idx]
            # rowid is 1-based source-file row number
            # Source row number = offset + i + 1 (since offset is 0-based row skip, and i is row index in paginated result)
            obj['rowid'] = offset + i + 1
            formatted_rows.append(obj)
    else:
        formatted_rows = paginated_rows

    # Validate visibility toggles
    show_rowid = True
    if '_rowid' in args:
        rowid_val = args['_rowid']
        if rowid_val != 'hide':
            return json_error(400, f"Invalid '_rowid' value: {rowid_val}")
        show_rowid = False

    show_total = True
    if '_total' in args:
        total_val = args['_total']
        if total_val != 'hide':
            return json_error(400, f"Invalid '_total' value: {total_val}")
        show_total = False

    # Check for query timeout
    if time.time() - start_time > 30:
        return json_error(400, "Query timeout")

    query_ms = (time.time() - start_time) * 1000

    result = {
        "ok": True,
        "columns": columns,
        "rows": formatted_rows,
        "query_ms": round(query_ms, 1)
    }

    if show_total:
        result['total'] = total

    # If shape=objects and rowid should be hidden, remove it from the response
    if shape == 'objects' and not show_rowid:
        for row_obj in result['rows']:
            del row_obj['rowid']

    return json_success(200, result)


@app.route('/datasets/<dataset_id>/export', methods=['GET'])
def export_dataset(dataset_id):
    """Export dataset as CSV with filters, sort, and pagination applied.

    Does not apply _shape, _rowid, or _total controls.
    Returns CSV bytes with Content-Type: text/csv and Content-Disposition header.
    """
    if dataset_id not in datasets:
        return json_error(404, "Dataset not found")

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = [list(row) for row in dataset['rows']]  # Create mutable copy
    args = request.args

    # Skip export-specific controls
    export_args = {k: v for k, v in args.items() if k not in ('_shape', '_rowid', '_total')}

    # Parse and validate duplicate parameters
    for key in export_args.keys():
        values = request.args.getlist(key)
        if len(values) > 1:
            if key.startswith('_'):
                return json_error(400, f"Repeated parameter: {key}")
            else:
                return json_error(400, f"Duplicate filter key: {key}")

    # --- Filter parsing and validation ---
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filter_params = {}
    numeric_filter_values = {}

    for key in export_args.keys():
        if key.startswith('_'):
            continue
        if '__' in key:
            parts = key.split('__', 1)
            if len(parts) == 2:
                col_name, comparator = parts
                if comparator not in valid_comparators:
                    return json_error(400, f"Invalid comparator: {comparator}")
                if col_name not in columns:
                    return json_error(400, f"Unknown filter column: {col_name}")

                filter_value = request.args.getlist(key)[0]
                if comparator in ('less', 'greater'):
                    try:
                        numeric_filter_values[key] = float(filter_value)
                    except (ValueError, TypeError):
                        return json_error(400, f"Invalid numeric value: {filter_value}")

                filter_params[key] = filter_value

    # --- Apply filters ---
    def apply_filters(row):
        for filter_key, filter_value in filter_params.items():
            col_name, comparator = filter_key.split('__', 1)
            col_idx = columns.index(col_name)
            cell_value = row[col_idx]

            if comparator == 'exact':
                if str(cell_value) != filter_value:
                    return False
            elif comparator == 'contains':
                if filter_value not in str(cell_value):
                    return False
            elif comparator == 'less' or comparator == 'greater':
                try:
                    cell_num = float(cell_value)
                except (ValueError, TypeError):
                    return False
                filter_num = numeric_filter_values[filter_key]
                if comparator == 'less':
                    if not (cell_num < filter_num):
                        return False
                elif comparator == 'greater':
                    if not (cell_num > filter_num):
                        return False
        return True

    filtered_rows = []
    for row in rows:
        if apply_filters(row):
            filtered_rows.append(row)
    rows = filtered_rows

    # Validate and get _size (default to all rows for export)
    size = len(rows)
    if '_size' in args:
        size_str = args['_size']
        if not size_str.isdigit() or int(size_str) <= 0:
            return json_error(400, f"Invalid '_size': {size_str}")
        size = int(size_str)

    # Validate and get _offset
    offset = 0
    if '_offset' in args:
        offset_str = args['_offset']
        if not offset_str.isdigit() or int(offset_str) < 0:
            return json_error(400, f"Invalid '_offset': {offset_str}")
        offset = int(offset_str)

    # Validate and get _sort/_sort_desc columns
    sort_col = args.get('_sort_desc') or args.get('_sort')
    sort_desc = '_sort_desc' in args

    if sort_col is not None:
        if sort_col == '' or sort_col not in columns:
            return json_error(400, f"Unknown column for sorting: {sort_col}")
        # Precompute column index mapping for O(1) lookups
        col_to_idx = {name: i for i, name in enumerate(columns)}
        col_idx = col_to_idx[sort_col]
        def sort_key(row):
            key = [row[col_idx]]
            for c in columns:
                if c != sort_col:
                    key.append(row[col_to_idx[c]])
            return tuple(key)
        rows.sort(key=sort_key, reverse=sort_desc)

    # Apply pagination
    paginated_rows = rows[offset:offset + size]

    # Generate CSV with source column order
    output = io.StringIO()
    writer = csv.writer(output)
    # Write header in original column order
    writer.writerow(columns)
    # Write data rows
    for row in paginated_rows:
        writer.writerow(row)

    csv_bytes = output.getvalue().encode('utf-8')

    response = Response(
        csv_bytes,
        status=200,
        mimetype='text/csv'
    )
    add_cors_headers(response)
    response.headers['Content-Disposition'] = f'attachment; filename="{dataset_id}.csv"'

    return response


def json_error(status_code: int, message: str) -> Response:
    response = Response(
        json.dumps({"ok": False, "error": message}),
        status=status_code,
        mimetype='application/json'
    )
    add_cors_headers(response)
    return response


def json_success(status_code: int, data: dict) -> Response:
    response = Response(
        json.dumps(data),
        status=status_code,
        mimetype='application/json'
    )
    add_cors_headers(response)
    return response


def add_cors_headers(response: Response):
    """Add CORS headers to response."""
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'


@app.route('/<path:unknown_path>', methods=['GET', 'OPTIONS'])
def catch_all(unknown_path):
    """Handle unknown routes with 404."""
    return json_error(404, "Not found")


@app.before_request
def handle_options():
    """Handle OPTIONS requests for CORS preflight."""
    if request.method == 'OPTIONS':
        response = Response()
        add_cors_headers(response)
        return response


def main():
    parser = argparse.ArgumentParser(description='datagate CSV conversion service')
    parser.add_argument('command', choices=['start'])
    parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')

    global datasets, url_to_id, file_to_id

    args = parser.parse_args()

    if args.command == 'start':
        # Ensure storage directory exists
        ensure_storage_dir()

        # Load persisted datasets at startup
        loaded_datasets, loaded_url_to_id, loaded_file_to_id = load_all_datasets()
        datasets.update(loaded_datasets)
        url_to_id.update(loaded_url_to_id)
        file_to_id.update(loaded_file_to_id)

        print(f"Loaded {len(datasets)} persisted datasets")

        app.run(host=args.address, port=args.port, debug=False)


if __name__ == '__main__':
    main()
