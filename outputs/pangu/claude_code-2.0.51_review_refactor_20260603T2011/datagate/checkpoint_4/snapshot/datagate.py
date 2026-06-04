#!/usr/bin/env python3
"""datagate - A CSV conversion and dataset query service."""

import argparse
import csv
import hashlib
import io
import json
import re
import time
from urllib.parse import urlparse

import chardet
import requests
from flask import Flask, Response, request

# For Excel support - optional dependency
try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# For xlrd support for .xls files
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# Pre-compiled regex for time pattern matching
_TIME_PATTERN = re.compile(r'^\d{1,2}:\d{2}$')

app = Flask(__name__)

# Storage for datasets: id -> data
datasets = {}

# Map source URL to dataset id for determinism
url_to_id = {}

# Map file content hash to dataset id for determinism (for upload endpoint)
file_to_id = {}


def generate_id(source_url: str) -> str:
    return hashlib.md5(source_url.encode('utf-8')).hexdigest()[:12]


def generate_file_id(file_bytes: bytes) -> str:
    return hashlib.md5(file_bytes).hexdigest()[:12]


def infer_delimiter(sample: str) -> str:
    """Infer CSV delimiter from sample content."""
    delimiters = [',', '\t', ';']
    sniffer = csv.Sniffer()

    # Try sniffer first
    try:
        dialect = sniffer.sniff(sample)
        return dialect.delimiter
    except csv.Error:
        pass

    # Manual detection
    best_delim = ','
    best_count = 0

    for delim in delimiters:
        count = sample.count(delim)
        if count > best_count:
            best_count = count
            best_delim = delim

    return best_delim


def is_time_like(value: str) -> bool:
    """Check if a string value looks like a time (e.g., 08:30, 9:15, 12:00)."""
    return bool(_TIME_PATTERN.match(value.strip()))


def convert_value(value: str):
    value = value.strip()

    if is_time_like(value):
        return value

    # Try integer
    try:
        return int(value)
    except ValueError:
        pass

    # Try float/decimal
    try:
        return float(value)
    except ValueError:
        pass

    # Return as string
    return value


def parse_csv(content: str, delimiter: str) -> dict:
    """Parse CSV content and return structured data."""
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)

    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("CSV must have at least one header row and one data row")

    columns = rows[0]
    data_rows = rows[1:]

    # Convert each row with type inference
    converted_rows = []
    for row in data_rows:
        converted_row = [convert_value(cell) for cell in row]
        converted_rows.append(converted_row)

    return {
        'columns': columns,
        'rows': converted_rows
    }


def fetch_csv(source_url: str, charset: str = None) -> tuple:
    """Fetch CSV from remote URL and return (raw_content, encoding)."""
    try:
        response = requests.get(source_url, timeout=30)
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}")
    except Exception as e:
        raise

    content_bytes = response.content

    if not content_bytes.strip():
        raise ValueError("Empty content")

    # Determine encoding
    if charset:
        try:
            content = content_bytes.decode(charset)
            encoding = charset
        except (UnicodeDecodeError, LookupError):
            raise ValueError(f"Invalid charset: {charset}")
    else:
        detected = chardet.detect(content_bytes)
        encoding = detected.get('encoding', 'utf-8')
        try:
            content = content_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            # Fallback to UTF-8
            encoding = 'utf-8'
            content = content_bytes.decode('utf-8', errors='replace')

    return content, encoding


def detect_file_type(file_bytes: bytes) -> str:
    """Detect file type from magic bytes.

    Returns: 'csv', 'xls', 'xlsx', or None if unknown.
    """
    if not file_bytes:
        return None

    # Check for ZIP (xlsx files are ZIP archives)
    if file_bytes[:2] == b'PK':
        # xlsx files start with PK ( ZIP format)
        return 'xlsx'

    # Check for .xls (OLE2 Compound Document)
    if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return 'xls'

    # For CSV, try to decode as text
    try:
        # Try UTF-8 with BOM
        if file_bytes[:3] == b'\xef\xbb\xbf':
            text = file_bytes[3:].decode('utf-8')
        else:
            text = file_bytes.decode('utf-8')

        # Check if it looks like a CSV (at least one comma or tab/semicolon)
        lines = text.strip().split('\n')
        if len(lines) >= 2:
            header = lines[0]
            if ',' in header or '\t' in header or ';' in header:
                # Verify data rows have consistent structure
                delimiters = [',', '\t', ';']
                for delim in delimiters:
                    if delim in header:
                        fields_in_header = header.split(delim)
                        # Check first few data rows
                        for line in lines[1:4]:
                            if line.strip():
                                fields = line.split(delim)
                                if len(fields) != len(fields_in_header):
                                    break
                        else:
                            return 'csv'
                        break
    except (UnicodeDecodeError, UnicodeError):
        pass

    return None


def parse_excel_xlsx(file_bytes: bytes) -> dict:
    """Parse .xlsx Excel file and return structured data."""
    if not HAS_OPENPYXL:
        raise ValueError("openpyxl not installed")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Get all rows from the first worksheet
    rows_data = []
    for row in ws.iter_rows(values_only=True):
        rows_data.append(list(row))

    if len(rows_data) < 2:
        raise ValueError("Excel must have at least one header row and one data row")

    columns = [str(c) if c is not None else "" for c in rows_data[0]]
    data_rows = rows_data[1:]

    # Convert each row with type inference
    converted_rows = []
    for row in data_rows:
        converted_row = []
        for cell in row:
            if cell is None:
                converted_row.append("")
            else:
                # Convert value with type inference
                converted_row.append(convert_value(str(cell)))
        converted_rows.append(converted_row)

    return {
        'columns': columns,
        'rows': converted_rows
    }


def parse_excel_xls(file_bytes: bytes) -> dict:
    """Parse .xls Excel file and return structured data."""
    if not HAS_XLRD:
        raise ValueError("xlrd not installed")

    book = xlrd.open_workbook(file_contents=file_bytes)
    ws = book.sheet_by_index(0)

    if ws.nrows < 2:
        raise ValueError("Excel must have at least one header row and one data row")

    # Get header row
    columns = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]
    data_rows = []

    # Get data rows
    for row_idx in range(1, ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell_value(row_idx, col_idx)
            row.append(convert_value(str(cell)))
        data_rows.append(row)

    return {
        'columns': columns,
        'rows': data_rows
    }


def is_tabular(content: str) -> bool:
    """Check if content appears to be tabular data."""
    lines = content.strip().split('\n')
    if len(lines) < 2:
        return False

    # Check that first line (header) has multiple fields
    header = lines[0]
    delimiters = [',', '\t', ';']

    for delim in delimiters:
        if delim in header:
            # Verify data rows have similar structure
            fields_in_header = header.split(delim)
            for line in lines[1:4]:  # Check first few data rows
                if line.strip():
                    fields = line.split(delim)
                    if len(fields) != len(fields_in_header):
                        return False
            return True

    return False


@app.route('/convert')
def convert():
    """Convert remote file to dataset.

    Supports CSV, .xls, and .xlsx from URLs.
    charset applies only to text CSV sources.
    """
    source = request.args.get('source')
    charset = request.args.get('charset')

    if not source:
        return json_error(400, "Missing 'source' parameter")

    # Validate URL
    try:
        parsed = urlparse(source)
        if not parsed.scheme or not parsed.netloc:
            raise ValueError("Invalid URL")
    except Exception:
        return json_error(400, "Invalid URL")

    if charset is not None and charset.strip() == '':
        return json_error(400, "Invalid charset: empty string")

    # Check if we already have this dataset
    if source in url_to_id:
        dataset_id = url_to_id[source]
        return json_success(200, {"ok": True, "endpoint": f"/datasets/{dataset_id}"})

    # Fetch the file - try as binary first to detect format
    try:
        response = requests.get(source, timeout=30)
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}")
        file_bytes = response.content
    except Exception as e:
        error_msg = str(e)
        # Remote HTTP 404 should return 404
        if "HTTP 404" in error_msg:
            return json_error(404, "Source unreachable")
        # Other errors are bad requests
        return json_error(400, error_msg)

    if not file_bytes.strip():
        return json_error(400, "Empty content")

    # Detect file type from magic bytes
    file_type = detect_file_type(file_bytes)

    if file_type is None:
        return json_error(400, "Unsupported format")

    # For CSV, charset applies
    if file_type == 'csv':
        # Determine encoding
        if charset:
            try:
                content = file_bytes.decode(charset)
                encoding = charset
            except (UnicodeDecodeError, LookupError):
                return json_error(400, f"Invalid charset: {charset}")
        else:
            detected = chardet.detect(file_bytes)
            encoding = detected.get('encoding', 'utf-8')
            try:
                content = file_bytes.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                encoding = 'utf-8'
                content = file_bytes.decode('utf-8', errors='replace')

        if not is_tabular(content):
            return json_error(400, "Non-tabular content")

        sample = content[:4096]
        delimiter = infer_delimiter(sample)

        try:
            dataset = parse_csv(content, delimiter)
        except ValueError as e:
            return json_error(400, str(e))

    elif file_type == 'xlsx':
        if not HAS_OPENPYXL:
            return json_error(400, "Unsupported format: .xlsx requires openpyxl")
        try:
            dataset = parse_excel_xlsx(file_bytes)
        except ValueError as e:
            return json_error(400, str(e))
        except Exception as e:
            return json_error(400, f"Invalid file: {str(e)}")

    elif file_type == 'xls':
        if not HAS_XLRD:
            return json_error(400, "Unsupported format: .xls requires xlrd")
        try:
            dataset = parse_excel_xls(file_bytes)
        except ValueError as e:
            return json_error(400, str(e))
        except Exception as e:
            return json_error(400, f"Invalid file: {str(e)}")

    else:
        return json_error(400, "Unsupported format")

    # Generate deterministic ID
    dataset_id = generate_id(source)

    # Store dataset
    datasets[dataset_id] = dataset
    url_to_id[source] = dataset_id

    return json_success(200, {"ok": True, "endpoint": f"/datasets/{dataset_id}"})


@app.route('/upload', methods=['POST'])
def upload():
    """Upload a file for conversion.

    Accepts multipart form with field 'file' or 'attachment'.
    Supports CSV, .xls, and .xlsx formats.
    """
    # Check for file field (in Flask, request.files exists even for non-multipart)
    # But we need to verify it's actually a multipart request
    # In Flask, request.files is only populated for multipart/form-data
    # If we get here but there's no files dict entry, it wasn't multipart
    # We check if we can access a file from the form

    # Check if files dict has any entries (multipart)
    # or try to get a file using .get() which returns None if not multipart
    try:
        uploaded_file = request.files.get('file') or request.files.get('attachment')
        if uploaded_file is None:
            # Check if files dict exists and is empty vs request isn't multipart
            # In Flask, request.files always exists for any request, but is empty for non-multipart
            # Also check content-type for safety
            content_type = request.content_type or ''
            if 'multipart' not in content_type.lower():
                return json_error(415, "Not a multipart request")
    except RuntimeError:
        return json_error(415, "Not a multipart request")

    if not uploaded_file:
        return json_error(400, "Missing 'file' field")

    # Read file content
    file_bytes = uploaded_file.read()

    if not file_bytes:
        return json_error(400, "Empty file")

    # Detect file type
    file_type = detect_file_type(file_bytes)

    if file_type is None:
        return json_error(400, "Unsupported format")

    # Check for unsupported formats
    if file_type == 'xls' and not HAS_XLRD:
        return json_error(400, "Unsupported format: .xls requires xlrd")
    if file_type == 'xlsx' and not HAS_OPENPYXL:
        return json_error(400, "Unsupported format: .xlsx requires openpyxl")

    # For CSV, we also accept without openpyxl or xlrd
    if file_type not in ('csv', 'xls', 'xlsx'):
        return json_error(400, "Unsupported format")

    # Generate deterministic ID based on file content
    dataset_id = generate_file_id(file_bytes)

    # Check if we already have this dataset
    if dataset_id in datasets:
        return json_success(200, {"ok": True, "endpoint": f"/datasets/{dataset_id}"})

    # Parse the file
    try:
        if file_type == 'csv':
            # Try to decode as UTF-8
            try:
                text = file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                # Try chardet for encoding detection
                detected = chardet.detect(file_bytes)
                encoding = detected.get('encoding', 'utf-8')
                text = file_bytes.decode(encoding, errors='replace')

            if not is_tabular(text):
                return json_error(400, "Non-tabular content")

            sample = text[:4096]
            delimiter = infer_delimiter(sample)
            dataset = parse_csv(text, delimiter)
        elif file_type == 'xlsx':
            dataset = parse_excel_xlsx(file_bytes)
        elif file_type == 'xls':
            dataset = parse_excel_xls(file_bytes)
        else:
            return json_error(400, "Unsupported format")
    except ValueError as e:
        return json_error(400, str(e))
    except Exception as e:
        return json_error(400, f"Invalid file: {str(e)}")

    # Store dataset
    datasets[dataset_id] = dataset
    file_to_id[dataset_id] = dataset_id

    return json_success(200, {"ok": True, "endpoint": f"/datasets/{dataset_id}"})


@app.route('/convert', methods=['POST'])
def convert_post():
    """Convert remote CSV to dataset (POST version).

    Accepts multipart form with field 'url' or 'source' for remote URL,
    or field 'file' or 'attachment' for direct file upload.
    """
    # Check if there's a file upload
    if request.is_multipart:
        uploaded_file = request.files.get('file') or request.files.get('attachment')
        if uploaded_file:
            # Use the upload endpoint logic but return from here
            return upload()

    # Check for URL parameter (can be form field or query string)
    source = request.form.get('url') or request.form.get('source') or request.args.get('source')
    charset = request.form.get('charset') or request.args.get('charset')

    if not source:
        return json_error(400, "Missing 'source' parameter")

    # Validate URL
    try:
        parsed = urlparse(source)
        if not parsed.scheme or not parsed.netloc:
            raise ValueError("Invalid URL")
    except Exception:
        return json_error(400, "Invalid URL")

    if charset is not None and charset.strip() == '':
        return json_error(400, "Invalid charset: empty string")

    # Check if we already have this dataset
    if source in url_to_id:
        dataset_id = url_to_id[source]
        return json_success(200, {"ok": True, "endpoint": f"/datasets/{dataset_id}"})

    # Fetch the CSV
    try:
        content, encoding = fetch_csv(source, charset)
    except Exception as e:
        error_msg = str(e)
        # Remote HTTP 404 should return 404
        if "HTTP 404" in error_msg:
            return json_error(404, "Source unreachable")
        # Other errors are bad requests
        return json_error(400, error_msg)

    # Validate it's tabular
    if not is_tabular(content):
        return json_error(400, "Non-tabular content")

    # Infer delimiter and parse
    sample = content[:4096]  # Sample for delimiter detection
    delimiter = infer_delimiter(sample)

    try:
        dataset = parse_csv(content, delimiter)
    except ValueError as e:
        return json_error(400, str(e))

    # Generate deterministic ID
    dataset_id = generate_id(source)

    # Store dataset
    datasets[dataset_id] = dataset
    url_to_id[source] = dataset_id

    return json_success(200, {"ok": True, "endpoint": f"/datasets/{dataset_id}"})


@app.route('/datasets/<dataset_id>')
def get_dataset(dataset_id):
    """Retrieve dataset by ID with pagination, sorting, and response controls."""
    if dataset_id not in datasets:
        return json_error(404, "Dataset not found")

    start_time = time.time()

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = [list(row) for row in dataset['rows']]  # Create mutable copy

    # Parse control parameters
    args = request.args

    # Check for duplicate parameters (control and filter) using getlist
    for key in args.keys():
        values = request.args.getlist(key)
        if len(values) > 1:
            if key.startswith('_'):
                return json_error(400, f"Repeated parameter: {key}")
            else:
                # It's a filter key
                return json_error(400, f"Duplicate filter key: {key}")

    # --- Filter parsing and validation ---
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filter_params = {}
    numeric_filter_values = {}  # Store parsed numeric values for less/greater

    for key in args.keys():
        # Skip control params (start with _)
        if key.startswith('_'):
            continue
        # Check for column__comparator pattern
        if '__' in key:
            parts = key.split('__', 1)
            if len(parts) == 2:
                col_name, comparator = parts
                if comparator not in valid_comparators:
                    return json_error(400, f"Invalid comparator: {comparator}")
                if col_name not in columns:
                    return json_error(400, f"Unknown filter column: {col_name}")

                filter_value = request.args.getlist(key)[0]
                # Validate numeric filter values for less/greater
                if comparator in ('less', 'greater'):
                    try:
                        numeric_filter_values[key] = float(filter_value)
                    except (ValueError, TypeError):
                        return json_error(400, f"Invalid numeric value: {filter_value}")

                filter_params[key] = filter_value

    # --- Apply filters ---
    def apply_filters(row):
        for filter_key, filter_value in filter_params.items():
            col_name, comparator = filter_key.split('__', 1)
            col_idx = columns.index(col_name)
            cell_value = row[col_idx]

            if comparator == 'exact':
                if str(cell_value) != filter_value:
                    return False
            elif comparator == 'contains':
                if filter_value not in str(cell_value):
                    return False
            elif comparator == 'less' or comparator == 'greater':
                # Skip rows with non-numeric stored values
                try:
                    cell_num = float(cell_value)
                except (ValueError, TypeError):
                    return False
                filter_num = numeric_filter_values[filter_key]
                if comparator == 'less':
                    if not (cell_num < filter_num):
                        return False
                elif comparator == 'greater':
                    if not (cell_num > filter_num):
                        return False
        return True

    # Apply filters to rows
    filtered_rows = []
    for row in rows:
        if apply_filters(row):
            filtered_rows.append(row)

    rows = filtered_rows

    # Validate and get _size
    size = 100
    if '_size' in args:
        size_str = args['_size']
        if not size_str.isdigit() or int(size_str) <= 0:
            return json_error(400, f"Invalid '_size': {size_str}")
        size = int(size_str)

    # Validate and get _offset
    offset = 0
    if '_offset' in args:
        offset_str = args['_offset']
        if not offset_str.isdigit() or int(offset_str) < 0:
            return json_error(400, f"Invalid '_offset': {offset_str}")
        offset = int(offset_str)

    # Validate and get _shape
    shape = 'lists'
    if '_shape' in args:
        shape = args['_shape']
        if shape not in ['lists', 'objects']:
            return json_error(400, f"Invalid '_shape': {shape}")

    # Validate _sort/_sort_desc columns and get sort parameters
    sort_col = args.get('_sort_desc') or args.get('_sort')
    sort_desc = '_sort_desc' in args

    # Validate sort column
    if sort_col is not None:
        if sort_col == '' or sort_col not in columns:
            return json_error(400, f"Unknown column for sorting: {sort_col}")
        col_idx = columns.index(sort_col)
        # Stable sort: for each row, use the sort column value, and for tie-breaking,
        # use the values of all columns in order
        def sort_key(row):
            key = [row[col_idx]]
            for c in columns:
                if c != sort_col:
                    key.append(row[columns.index(c)])
            return tuple(key)
        rows.sort(key=sort_key, reverse=sort_desc)

    total = len(rows)

    # Apply pagination
    paginated_rows = rows[offset:offset + size]

    # Format rows based on shape
    if shape == 'objects':
        formatted_rows = []
        for i, row in enumerate(paginated_rows):
            obj = {}
            for col_idx, col_name in enumerate(columns):
                obj[col_name] = row[col_idx]
            # rowid is 1-based source-file row number
            # Source row number = offset + i + 1 (since offset is 0-based row skip, and i is row index in paginated result)
            obj['rowid'] = offset + i + 1
            formatted_rows.append(obj)
    else:
        formatted_rows = paginated_rows

    # Validate visibility toggles
    show_rowid = True
    if '_rowid' in args:
        rowid_val = args['_rowid']
        if rowid_val != 'hide':
            return json_error(400, f"Invalid '_rowid' value: {rowid_val}")
        show_rowid = False

    show_total = True
    if '_total' in args:
        total_val = args['_total']
        if total_val != 'hide':
            return json_error(400, f"Invalid '_total' value: {total_val}")
        show_total = False

    # Check for query timeout
    if time.time() - start_time > 30:
        return json_error(400, "Query timeout")

    query_ms = (time.time() - start_time) * 1000

    result = {
        "ok": True,
        "columns": columns,
        "rows": formatted_rows,
        "query_ms": round(query_ms, 1)
    }

    if show_total:
        result['total'] = total

    # If shape=objects and rowid should be hidden, remove it from the response
    if shape == 'objects' and not show_rowid:
        for row_obj in result['rows']:
            del row_obj['rowid']

    return json_success(200, result)


@app.route('/datasets/<dataset_id>/export', methods=['GET'])
def export_dataset(dataset_id):
    """Export dataset as CSV with filters, sort, and pagination applied.

    Does not apply _shape, _rowid, or _total controls.
    Returns CSV bytes with Content-Type: text/csv and Content-Disposition header.
    """
    if dataset_id not in datasets:
        return json_error(404, "Dataset not found")

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = [list(row) for row in dataset['rows']]  # Create mutable copy
    args = request.args

    # Skip export-specific controls
    export_args = {k: v for k, v in args.items() if k not in ('_shape', '_rowid', '_total')}

    # Parse and validate duplicate parameters
    for key in export_args.keys():
        values = request.args.getlist(key)
        if len(values) > 1:
            if key.startswith('_'):
                return json_error(400, f"Repeated parameter: {key}")
            else:
                return json_error(400, f"Duplicate filter key: {key}")

    # --- Filter parsing and validation ---
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filter_params = {}
    numeric_filter_values = {}

    for key in export_args.keys():
        if key.startswith('_'):
            continue
        if '__' in key:
            parts = key.split('__', 1)
            if len(parts) == 2:
                col_name, comparator = parts
                if comparator not in valid_comparators:
                    return json_error(400, f"Invalid comparator: {comparator}")
                if col_name not in columns:
                    return json_error(400, f"Unknown filter column: {col_name}")

                filter_value = request.args.getlist(key)[0]
                if comparator in ('less', 'greater'):
                    try:
                        numeric_filter_values[key] = float(filter_value)
                    except (ValueError, TypeError):
                        return json_error(400, f"Invalid numeric value: {filter_value}")

                filter_params[key] = filter_value

    # --- Apply filters ---
    def apply_filters(row):
        for filter_key, filter_value in filter_params.items():
            col_name, comparator = filter_key.split('__', 1)
            col_idx = columns.index(col_name)
            cell_value = row[col_idx]

            if comparator == 'exact':
                if str(cell_value) != filter_value:
                    return False
            elif comparator == 'contains':
                if filter_value not in str(cell_value):
                    return False
            elif comparator == 'less' or comparator == 'greater':
                try:
                    cell_num = float(cell_value)
                except (ValueError, TypeError):
                    return False
                filter_num = numeric_filter_values[filter_key]
                if comparator == 'less':
                    if not (cell_num < filter_num):
                        return False
                elif comparator == 'greater':
                    if not (cell_num > filter_num):
                        return False
        return True

    filtered_rows = []
    for row in rows:
        if apply_filters(row):
            filtered_rows.append(row)
    rows = filtered_rows

    # Validate and get _size (default to all rows for export)
    size = len(rows)
    if '_size' in args:
        size_str = args['_size']
        if not size_str.isdigit() or int(size_str) <= 0:
            return json_error(400, f"Invalid '_size': {size_str}")
        size = int(size_str)

    # Validate and get _offset
    offset = 0
    if '_offset' in args:
        offset_str = args['_offset']
        if not offset_str.isdigit() or int(offset_str) < 0:
            return json_error(400, f"Invalid '_offset': {offset_str}")
        offset = int(offset_str)

    # Validate and get _sort/_sort_desc columns
    sort_col = args.get('_sort_desc') or args.get('_sort')
    sort_desc = '_sort_desc' in args

    if sort_col is not None:
        if sort_col == '' or sort_col not in columns:
            return json_error(400, f"Unknown column for sorting: {sort_col}")
        col_idx = columns.index(sort_col)
        def sort_key(row):
            key = [row[col_idx]]
            for c in columns:
                if c != sort_col:
                    key.append(row[columns.index(c)])
            return tuple(key)
        rows.sort(key=sort_key, reverse=sort_desc)

    # Apply pagination
    paginated_rows = rows[offset:offset + size]

    # Generate CSV with source column order
    output = io.StringIO()
    writer = csv.writer(output)
    # Write header in original column order
    writer.writerow(columns)
    # Write data rows
    for row in paginated_rows:
        writer.writerow(row)

    csv_bytes = output.getvalue().encode('utf-8')

    response = Response(
        csv_bytes,
        status=200,
        mimetype='text/csv'
    )
    add_cors_headers(response)
    response.headers['Content-Disposition'] = f'attachment; filename="{dataset_id}.csv"'

    return response


def json_error(status_code: int, message: str) -> Response:
    response = Response(
        json.dumps({"ok": False, "error": message}),
        status=status_code,
        mimetype='application/json'
    )
    add_cors_headers(response)
    return response


def json_success(status_code: int, data: dict) -> Response:
    response = Response(
        json.dumps(data),
        status=status_code,
        mimetype='application/json'
    )
    add_cors_headers(response)
    return response


def add_cors_headers(response: Response):
    """Add CORS headers to response."""
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'


@app.route('/<path:unknown_path>', methods=['GET', 'OPTIONS'])
def catch_all(unknown_path):
    """Handle unknown routes with 404."""
    return json_error(404, "Not found")


@app.before_request
def handle_options():
    """Handle OPTIONS requests for CORS preflight."""
    if request.method == 'OPTIONS':
        response = Response()
        add_cors_headers(response)
        return response


def main():
    parser = argparse.ArgumentParser(description='datagate CSV conversion service')
    parser.add_argument('command', choices=['start'])
    parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')

    args = parser.parse_args()

    if args.command == 'start':
        app.run(host=args.address, port=args.port, debug=False)


if __name__ == '__main__':
    main()
