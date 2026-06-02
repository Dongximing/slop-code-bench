#!/usr/bin/env python3
"""
Test the new features: export, upload, and multi-format support.
"""

import csv
import json
import os
import subprocess
import sys
import tempfile
import time
import threading
from http.server import HTTPServer, SimpleHTTPRequestHandler

import requests

BASE_URL = 'http://127.0.0.1:8001'

def create_test_csv(content, delimiter=','):
    """Create a temporary CSV file and return a file URL and server."""
    fd, path = tempfile.mkstemp(suffix='.csv')
    with os.fdopen(fd, 'w', newline='') as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in content:
            writer.writerow(row)

    server_fd, server_path = tempfile.mkstemp()
    os.close(server_fd)
    os.unlink(server_path)

    class Handler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=os.path.dirname(path), **kwargs)

    server = HTTPServer(('127.0.0.1', 0), Handler)
    thread = threading.Thread(target=server.serve_forever)
    thread.daemon = True
    thread.start()

    host, port = server.server_address
    filename = os.path.basename(path)
    url = f'http://{host}:{port}/{filename}'

    time.sleep(0.5)
    return url, server


def test_export_basic():
    """Test basic CSV export."""
    print("Testing basic export...")
    csv_content = [
        ['name', 'age', 'score'],
        ['Alice', '25', '95.5'],
        ['Bob', '30', '87.2'],
        ['Charlie', '35', '92.0']
    ]
    url, server = create_test_csv(csv_content)
    try:
        r = requests.get(f'{BASE_URL}/convert', params={'source': url})
        endpoint = r.json()['endpoint']

        # Export as CSV
        r = requests.get(f'{BASE_URL}{endpoint}/export')
        assert r.status_code == 200
        assert r.headers['Content-Type'] == 'text/csv'
        assert 'attachment' in r.headers['Content-Disposition']

        # Verify CSV content
        lines = r.text.strip().split('\n')
        assert lines[0] == 'name,age,score'
        assert len(lines) == 4  # header + 3 rows

        # Verify values
        assert 'Alice,25,95.5' in r.text
        assert 'Bob,30,87.2' in r.text
        assert 'Charlie,35,92.0' in r.text
        print("  PASSED")
    finally:
        server.shutdown()


def test_export_with_filters():
    """Test export with filters."""
    print("Testing export with filters...")
    csv_content = [
        ['name', 'age'],
        ['Alice', '25'],
        ['Bob', '30'],
        ['Charlie', '35']
    ]
    url, server = create_test_csv(csv_content)
    try:
        r = requests.get(f'{BASE_URL}/convert', params={'source': url})
        endpoint = r.json()['endpoint']

        # Export with filter
        r = requests.get(f'{BASE_URL}{endpoint}/export', params={'age__exact': '30'})
        assert r.status_code == 200

        lines = r.text.strip().split('\n')
        assert len(lines) == 2  # header + 1 filtered row
        assert 'Bob,30' in r.text
        assert 'Alice,25' not in r.text
        print("  PASSED")
    finally:
        server.shutdown()


def test_export_with_sort():
    """Test export with sorting."""
    print("Testing export with sorting...")
    csv_content = [
        ['name'],
        ['Charlie'],
        ['Alice'],
        ['Bob']
    ]
    url, server = create_test_csv(csv_content)
    try:
        r = requests.get(f'{BASE_URL}/convert', params={'source': url})
        endpoint = r.json()['endpoint']

        # Export with sort
        r = requests.get(f'{BASE_URL}{endpoint}/export', params={'_sort': 'name'})
        assert r.status_code == 200

        lines = r.text.strip().split('\n')
        assert lines[1] == 'Alice'
        assert lines[2] == 'Bob'
        assert lines[3] == 'Charlie'
        print("  PASSED")
    finally:
        server.shutdown()


def test_export_with_pagination():
    """Test export with pagination parameters."""
    print("Testing export with pagination...")
    csv_content = [['num']] + [[str(i)] for i in range(10)]
    url, server = create_test_csv(csv_content)
    try:
        r = requests.get(f'{BASE_URL}/convert', params={'source': url})
        endpoint = r.json()['endpoint']

        # Export with _size
        r = requests.get(f'{BASE_URL}{endpoint}/export', params={'_size': '5'})
        assert r.status_code == 200
        lines = r.text.strip().split('\n')
        assert len(lines) == 6  # header + 5 rows

        # Export with _offset and _size
        r = requests.get(f'{BASE_URL}{endpoint}/export', params={'_offset': '5', '_size': '3'})
        assert r.status_code == 200
        lines = r.text.strip().split('\n')
        assert len(lines) == 4  # header + 3 rows
        assert lines[1] == '5'
        assert lines[2] == '6'
        assert lines[3] == '7'
        print("  PASSED")
    finally:
        server.shutdown()


def test_upload_csv():
    """Test CSV file upload."""
    print("Testing CSV upload...")
    csv_file = tempfile.NamedTemporaryFile(suffix='.csv', mode='w', newline='', delete=False)
    writer = csv.writer(csv_file)
    writer.writerow(['name', 'value'])
    writer.writerow(['test1', '100'])
    writer.writerow(['test2', '200'])
    csv_file.close()

    try:
        with open(csv_file.name, 'rb') as f:
            r = requests.post(f'{BASE_URL}/upload', files={'file': f})

        assert r.status_code == 200
        data = r.json()
        assert data.get('ok') is True
        assert 'endpoint' in data
        print("  PASSED")
    finally:
        os.unlink(csv_file.name)


def test_upload_xlsx():
    """Test XLSX file upload."""
    print("Testing XLSX upload...")
    try:
        import openpyxl
    except ImportError:
        print("  SKIPPED (openpyxl not available)")
        return

    # Create a temporary xlsx file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['name', 'value'])
    ws.append(['xlsx1', '100'])
    ws.append(['xlsx2', '200'])

    fd, path = tempfile.mkstemp(suffix='.xlsx')
    wb.save(path)

    try:
        with open(path, 'rb') as f:
            r = requests.post(f'{BASE_URL}/upload', files={'file': f})

        assert r.status_code == 200
        data = r.json()
        assert data.get('ok') is True
        assert 'endpoint' in data
        print("  PASSED")
    finally:
        os.unlink(path)


def test_upload_xls():
    """Test XLS file upload."""
    print("Testing XLS upload...")
    try:
        import xlrd
        print("  SKIPPED (xlrd doesn't support writing, skipping)")
        return
    except ImportError:
        print("  SKIPPED (xlrd not available)")
        return


def test_upload_xls_with_xlwt():
    """Test XLS file using xlwt if available."""
    print("Testing XLS upload...")
    try:
        import xlwt
    except ImportError:
        print("  SKIPPED (xlwt not available)")
        return

    # Create a temporary xls file
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')
    ws.write(0, 0, 'name')
    ws.write(0, 1, 'value')
    ws.write(1, 0, 'xls1')
    ws.write(1, 1, '300')
    ws.write(2, 0, 'xls2')
    ws.write(2, 1, '400')

    fd, path = tempfile.mkstemp(suffix='.xls')
    wb.save(path)

    try:
        with open(path, 'rb') as f:
            r = requests.post(f'{BASE_URL}/upload', files={'file': f})

        assert r.status_code == 200
        data = r.json()
        assert data.get('ok') is True
        assert 'endpoint' in data
        print("  PASSED")
    finally:
        os.unlink(path)


def test_upload_missing_field():
    """Test upload with missing file field."""
    print("Testing upload without file field...")

    # Try without file
    r = requests.post(f'{BASE_URL}/upload', data={'some_field': 'value'})
    assert r.status_code == 400
    data = r.json()
    assert data.get('ok') is False
    assert 'missing file field' in data.get('error', '').lower()
    print("  PASSED")


def test_upload_non_multipart():
    """Test non-multipart upload."""
    print("Testing non-multipart upload...")

    # Try form data (not multipart)
    r = requests.post(f'{BASE_URL}/upload', data={'file': 'content'})
    assert r.status_code == 415
    data = r.json()
    assert data.get('ok') is False
    assert 'non-multipart' in data.get('error', '').lower()
    print("  PASSED")


def test_upload_attachment_field():
    """Test upload using 'attachment' field."""
    print("Testing upload with 'attachment' field...")

    csv_file = tempfile.NamedTemporaryFile(suffix='.csv', mode='w', newline='', delete=False)
    writer = csv.writer(csv_file)
    writer.writerow(['x', 'y'])
    writer.writerow(['1', '2'])
    csv_file.close()

    try:
        with open(csv_file.name, 'rb') as f:
            r = requests.post(f'{BASE_URL}/upload', files={'attachment': f})

        assert r.status_code == 200
        data = r.json()
        assert data.get('ok') is True
        print("  PASSED")
    finally:
        os.unlink(csv_file.name)


def test_convert_xlsx_remote():
    """Test XLSX conversion from remote URL."""
    print("Testing XLSX conversion from remote URL...")
    try:
        import openpyxl
    except ImportError:
        print("  SKIPPED (openpyxl not available)")
        return

    # Create a temporary xlsx and serve it
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['a', 'b'])
    ws.append(['1', '2'])
    ws.append(['3', '4'])

    fd, path = tempfile.mkstemp(suffix='.xlsx')
    wb.save(path)

    class Handler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=os.path.dirname(path), **kwargs)

    server = HTTPServer(('127.0.0.1', 0), Handler)
    thread = threading.Thread(target=server.serve_forever)
    thread.daemon = True
    thread.start()

    host, port = server.server_address
    filename = os.path.basename(path)
    url = f'http://{host}:{port}/{filename}'
    time.sleep(0.5)

    try:
        r = requests.get(f'{BASE_URL}/convert', params={'source': url})
        assert r.status_code == 200
        data = r.json()
        assert data.get('ok') is True
        endpoint = data['endpoint']

        # Verify the data
        r = requests.get(f'{BASE_URL}{endpoint}')
        assert r.status_code == 200
        ds = r.json()
        assert ds['columns'] == ['a', 'b']
        assert len(ds['rows']) == 2
        print("  PASSED")
    finally:
        server.shutdown()
        os.unlink(path)


def test_upload_redeterminism():
    """Test that re-uploading same content yields same dataset ID."""
    print("Testing upload determinism...")

    csv_content = [['x', 'y'], ['1', '2']]

    # First upload
    fd1, path1 = tempfile.mkstemp(suffix='.csv')
    with os.fdopen(fd1, 'w', newline='') as f:
        writer = csv.writer(f)
        for row in csv_content:
            writer.writerow(row)

    with open(path1, 'rb') as f:
        r1 = requests.post(f'{BASE_URL}/upload', files={'file': f})
    endpoint1 = r1.json()['endpoint']

    # Second upload with same content
    time.sleep(0.5)
    fd2, path2 = tempfile.mkstemp(suffix='.csv')
    with os.fdopen(fd2, 'w', newline='') as f:
        writer = csv.writer(f)
        for row in csv_content:
            writer.writerow(row)

    with open(path2, 'rb') as f:
        r2 = requests.post(f'{BASE_URL}/upload', files={'file': f})
    endpoint2 = r2.json()['endpoint']

    assert endpoint1 == endpoint2
    print("  PASSED")

    os.unlink(path1)
    os.unlink(path2)


def main():
    # Test that server is running
    print("Waiting for server...")
    for _ in range(10):
        try:
            requests.get(f'{BASE_URL}/', timeout=1)
            break
        except requests.exceptions.ConnectionError:
            time.sleep(0.5)

    # Run all tests
    tests = [
        test_export_basic,
        test_export_with_filters,
        test_export_with_sort,
        test_export_with_pagination,
        test_upload_csv,
        test_upload_xlsx,
        test_upload_xls_with_xlwt,
        test_upload_missing_field,
        test_upload_non_multipart,
        test_upload_attachment_field,
        test_convert_xlsx_remote,
        test_upload_redeterminism,
    ]

    passed = 0
    failed = 0

    for test in tests:
        try:
            test()
            passed += 1
        except Exception as e:
            failed += 1
            print(f"  FAILED: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*50}")
    print(f"Results: {passed} passed, {failed} failed")
    return 0 if failed == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
