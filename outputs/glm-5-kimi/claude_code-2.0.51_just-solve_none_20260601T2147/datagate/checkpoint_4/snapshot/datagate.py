#!/usr/bin/env python3
"""datagate - A CSV data ingestion and query service."""

import argparse
import csv
import hashlib
import io
import json
import re
import time
import email.parser
import email.policy
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import urllib.request
import urllib.error
import charset_normalizer
import openpyxl
import xlrd


def _num_cmp(row_val, filter_val, op):
    try:
        return op(float(row_val), float(filter_val))
    except (ValueError, TypeError):
        return False


_COMPARATORS = {
    'exact': lambda v, f: str(v) == f,
    'contains': lambda v, f: f in str(v),
    'less': lambda v, f: _num_cmp(v, f, lambda a, b: a < b),
    'greater': lambda v, f: _num_cmp(v, f, lambda a, b: a > b),
}


def detect_delimiter(text: str) -> str:
    delimiters = [',', ';', '\t']
    first_line = text.split('\n')[0]
    counts = {d: first_line.count(d) for d in delimiters}
    best = max(delimiters, key=lambda d: counts[d])
    return best if counts[best] > 0 else ','


def is_time_like(value: str) -> bool:
    return bool(re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', value.strip()))


def parse_value(value: str):
    if value == '' or is_time_like(value):
        return value

    if '.' not in value and 'e' not in value.lower():
        try:
            return int(value)
        except ValueError:
            pass

    try:
        return float(value)
    except ValueError:
        return value


def generate_dataset_id(source_url: str) -> str:
    return hashlib.sha256(source_url.encode('utf-8')).hexdigest()[:16]


def is_valid_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
        return bool(parsed.scheme and parsed.netloc and parsed.scheme in ('http', 'https'))
    except Exception:
        return False


def validate_charset(charset: str) -> bool:
    try:
        'test'.encode(charset)
        return True
    except (LookupError, UnicodeError):
        return False


def decode_content(content: bytes, charset: str | None = None) -> str | None:
    if charset:
        try:
            return content.decode(charset)
        except UnicodeDecodeError:
            return None

    result = charset_normalizer.from_bytes(content)
    if result and (best := result.best()):
        return content.decode(best.encoding)

    for fallback in ('utf-8', 'latin-1'):
        try:
            return content.decode(fallback)
        except UnicodeDecodeError:
            continue
    return None


def _fetch_csv(source_url: str, charset: str | None = None) -> tuple[str, None] | tuple[None, str]:
    try:
        req = urllib.request.Request(source_url, headers={'User-Agent': 'datagate/1.0'})
        content = urllib.request.urlopen(req, timeout=30).read()
    except urllib.error.HTTPError as e:
        return None, f'404:Remote HTTP error: {e.code}'
    except urllib.error.URLError as e:
        return None, f'404:Source unreachable: {e.reason}'
    except Exception as e:
        return None, f'404:Source unreachable: {e}'

    text = decode_content(content, charset)
    if text is None:
        return None, f'400:{f"Failed to decode with charset: {charset}" if charset else "Could not detect encoding"}'
    return text, None


def _parse_csv(text: str) -> tuple[list, list, None] | tuple[None, None, str]:
    delimiter = detect_delimiter(text)
    try:
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    except Exception as e:
        return None, None, f'Failed to parse CSV: {e}'

    if len(rows) < 2:
        return None, None, 'Non-tabular content: requires at least one header row and one data row'

    num_cols = len(rows[0])
    if num_cols == 0:
        return None, None, 'Non-tabular content: no columns detected'
    if any(len(row) != num_cols for row in rows):
        return None, None, 'Non-tabular content: inconsistent column count'

    columns = rows[0]
    data_rows = [[parse_value(cell) for cell in row] for row in rows[1:]]
    return columns, data_rows, None


def _parse_xlsx(content: bytes) -> tuple[list, list, None] | tuple[None, None, str]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return None, None, f'Failed to parse XLSX: {e}'

    if not wb.sheetnames:
        wb.close()
        return None, None, 'Non-tabular content: no sheets found'

    ws = wb[wb.sheetnames[0]]
    raw_rows = []
    for row in ws.iter_rows():
        raw_rows.append([cell.value for cell in row])
    wb.close()

    return _validate_tabular_rows(raw_rows)


def _parse_xls(content: bytes) -> tuple[list, list, None] | tuple[None, None, str]:
    try:
        wb = xlrd.open_workbook(file_contents=content)
    except Exception as e:
        return None, None, f'Failed to parse XLS: {e}'

    if wb.nsheets == 0:
        return None, None, 'Non-tabular content: no sheets found'

    sheet = wb.sheet_by_index(0)
    if sheet.nrows < 2:
        return None, None, 'Non-tabular content: requires at least one header row and one data row'

    raw_rows = []
    for row_idx in range(sheet.nrows):
        raw_rows.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])

    return _validate_tabular_rows(raw_rows)


def _validate_tabular_rows(raw_rows: list) -> tuple[list, list, None] | tuple[None, None, str]:
    if len(raw_rows) < 2:
        return None, None, 'Non-tabular content: requires at least one header row and one data row'

    num_cols = len(raw_rows[0])
    if num_cols == 0:
        return None, None, 'Non-tabular content: no columns detected'
    if any(len(row) != num_cols for row in raw_rows):
        return None, None, 'Non-tabular content: inconsistent column count'

    columns = [str(c) if c is not None else '' for c in raw_rows[0]]
    data_rows = []
    for row in raw_rows[1:]:
        parsed = []
        for cell in row:
            if cell is None:
                parsed.append('')
            elif isinstance(cell, (int, float)):
                parsed.append(cell)
            else:
                parsed.append(parse_value(str(cell)))
        data_rows.append(parsed)
    return columns, data_rows, None


def _detect_format_from_bytes(content: bytes, filename: str | None = None) -> str | None:
    """Detect file format from content bytes and optional filename.

    Returns: 'csv', 'xlsx', 'xls', or None if unrecognized.
    """
    # Check by magic bytes first
    if content[:4] == b'PK\x03\x04':
        return 'xlsx'  # XLSX is a ZIP archive
    if content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return 'xls'  # OLE2 compound document

    # Check filename extension as fallback
    if filename:
        lower = filename.lower()
        if lower.endswith('.xlsx'):
            return 'xlsx'
        if lower.endswith('.xls'):
            return 'xls'
        if lower.endswith('.csv'):
            return 'csv'

    # Default to CSV for text content
    return 'csv'


def _parse_content(content: bytes, fmt: str, charset: str | None = None,
                   filename: str | None = None) -> tuple[list, list, None] | tuple[None, None, str]:
    """Parse content bytes into columns and rows based on format."""
    if fmt == 'xlsx':
        return _parse_xlsx(content)
    elif fmt == 'xls':
        return _parse_xls(content)
    elif fmt == 'csv':
        text = decode_content(content, charset)
        if text is None:
            return None, None, (f'Failed to decode with charset: {charset}'
                                if charset else 'Could not detect encoding')
        return _parse_csv(text)
    else:
        return None, None, f'Unsupported format: {fmt}'


_datasets: dict = {}


class DataGateHandler(BaseHTTPRequestHandler):
    def send_json_response(self, status_code: int, data: dict):
        self.send_response(status_code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode('utf-8'))

    def send_error_response(self, status_code: int, message: str):
        self.send_json_response(status_code, {'ok': False, 'error': message})

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        query_params = parse_qs(parsed_url.query)

        if path == '/convert':
            self.handle_convert(query_params)
        elif path.startswith('/datasets/'):
            remainder = path[len('/datasets/'):]
            if remainder.endswith('/export'):
                dataset_id = remainder[:-7]  # Remove '/export'
                self.handle_export(dataset_id)
            else:
                dataset_id = remainder
                self.handle_dataset(dataset_id)
        else:
            self.send_error_response(404, 'Not found')

    def do_POST(self):
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        query_params = parse_qs(parsed_url.query)

        if path == '/upload':
            self.handle_upload(query_params)
        elif path == '/convert':
            self.handle_convert_post(query_params)
        else:
            self.send_error_response(404, 'Not found')

    def handle_convert(self, query_params: dict):
        if 'source' not in query_params or not query_params['source']:
            self.send_error_response(400, 'Missing required parameter: source')
            return

        source = query_params['source'][0]

        if not is_valid_url(source):
            self.send_error_response(400, 'Invalid URL')
            return

        charset = query_params['charset'][0] if query_params.get('charset') else None
        if charset and not validate_charset(charset):
            self.send_error_response(400, f'Unsupported or malformed charset: {charset}')
            return

        dataset_id = generate_dataset_id(source)

        if dataset_id in _datasets:
            self.send_json_response(200, {
                'ok': True,
                'endpoint': f'/datasets/{dataset_id}'
            })
            return

        # Fetch content
        try:
            req = urllib.request.Request(source, headers={'User-Agent': 'datagate/1.0'})
            content = urllib.request.urlopen(req, timeout=30).read()
        except urllib.error.HTTPError as e:
            self.send_error_response(404, f'Remote HTTP error: {e.code}')
            return
        except urllib.error.URLError as e:
            self.send_error_response(404, f'Source unreachable: {e.reason}')
            return
        except Exception as e:
            self.send_error_response(404, f'Source unreachable: {e}')
            return

        # Detect format from URL and content
        filename = source.split('/')[-1] if '/' in source else source
        fmt = _detect_format_from_bytes(content, filename)

        columns, rows, err = _parse_content(content, fmt, charset, filename)
        if err:
            self.send_error_response(400, err)
            return

        _datasets[dataset_id] = {'columns': columns, 'rows': rows}

        self.send_json_response(200, {
            'ok': True,
            'endpoint': f'/datasets/{dataset_id}'
        })

    def _parse_multipart(self) -> tuple[bytes, str | None, str | None] | tuple[None, None, None]:
        """Parse multipart form data. Returns (content, filename, content_type) or (None, None, None) on error."""
        content_type = self.headers.get('Content-Type', '')
        if not content_type.startswith('multipart/form-data'):
            return None, None, None

        # Extract boundary
        boundary = None
        for part in content_type.split(';'):
            part = part.strip()
            if part.startswith('boundary='):
                boundary = part[9:]
                # Remove quotes if present
                if boundary.startswith('"') and boundary.endswith('"'):
                    boundary = boundary[1:-1]
                break

        if not boundary:
            return None, None, None

        # Read the request body
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)

        # Parse using email parser
        parser = email.parser.BytesParser(policy=email.policy.HTTP)
        msg = parser.parsebytes(b'Content-Type: ' + content_type.encode() + b'\r\n\r\n' + body)

        file_content = None
        filename = None
        file_content_type = None

        for part in msg.iter_parts():
            content_disposition = part.get('Content-Disposition', '')
            if 'form-data' not in content_disposition:
                continue

            # Extract field name and filename
            field_name = None
            fn = None
            for item in content_disposition.split(';'):
                item = item.strip()
                if item.startswith('name='):
                    field_name = item[5:].strip('"')
                elif item.startswith('filename='):
                    fn = item[9:].strip('"')

            # Check for file or attachment field
            if field_name in ('file', 'attachment'):
                file_content = part.get_payload(decode=True)
                filename = fn
                file_content_type = part.get_content_type()
                break

        return file_content, filename, file_content_type

    def handle_upload(self, query_params: dict):
        content_type = self.headers.get('Content-Type', '')

        # Check if multipart
        if not content_type.startswith('multipart/form-data'):
            self.send_error_response(415, 'Expected multipart/form-data')
            return

        # Parse multipart
        file_content, filename, file_ct = self._parse_multipart()

        if file_content is None:
            self.send_error_response(400, 'Missing file or attachment field')
            return

        # Get charset from query params (only applies to CSV)
        charset = query_params['charset'][0] if query_params.get('charset') else None
        if charset and not validate_charset(charset):
            self.send_error_response(400, f'Unsupported or malformed charset: {charset}')
            return

        # Generate dataset ID from file content hash
        dataset_id = hashlib.sha256(file_content).hexdigest()[:16]

        # Check if already exists
        if dataset_id in _datasets:
            self.send_json_response(200, {
                'ok': True,
                'endpoint': f'/datasets/{dataset_id}'
            })
            return

        # Detect format
        fmt = _detect_format_from_bytes(file_content, filename)

        # Parse content
        columns, rows, err = _parse_content(file_content, fmt, charset, filename)
        if err:
            self.send_error_response(400, err)
            return

        _datasets[dataset_id] = {'columns': columns, 'rows': rows}

        self.send_json_response(200, {
            'ok': True,
            'endpoint': f'/datasets/{dataset_id}'
        })

    def handle_convert_post(self, query_params: dict):
        content_type = self.headers.get('Content-Type', '')

        # Check if multipart
        if not content_type.startswith('multipart/form-data'):
            self.send_error_response(415, 'Expected multipart/form-data')
            return

        # Parse multipart
        file_content, filename, file_ct = self._parse_multipart()

        if file_content is None:
            self.send_error_response(400, 'Missing file or attachment field')
            return

        # Get charset from query params (only applies to CSV)
        charset = query_params['charset'][0] if query_params.get('charset') else None
        if charset and not validate_charset(charset):
            self.send_error_response(400, f'Unsupported or malformed charset: {charset}')
            return

        # Generate dataset ID from file content hash
        dataset_id = hashlib.sha256(file_content).hexdigest()[:16]

        # Check if already exists
        if dataset_id in _datasets:
            self.send_json_response(200, {
                'ok': True,
                'endpoint': f'/datasets/{dataset_id}'
            })
            return

        # Detect format
        fmt = _detect_format_from_bytes(file_content, filename)

        # Parse content
        columns, rows, err = _parse_content(file_content, fmt, charset, filename)
        if err:
            self.send_error_response(400, err)
            return

        _datasets[dataset_id] = {'columns': columns, 'rows': rows}

        self.send_json_response(200, {
            'ok': True,
            'endpoint': f'/datasets/{dataset_id}'
        })

    def handle_export(self, dataset_id: str):
        dataset = _datasets.get(dataset_id)
        if dataset is None:
            self.send_error_response(404, 'Dataset not found')
            return

        parsed_url = urlparse(self.path)
        query_params = parse_qs(parsed_url.query, keep_blank_values=True)

        # Parse filters and sort (same as handle_dataset but without pagination)
        columns = dataset['columns']
        rows = list(dataset['rows'])

        # Parse filters
        filters = self._parse_filters(query_params, columns)
        if filters is None:
            return

        rows = self._apply_filters(rows, columns, filters)

        # Parse sort
        sort_column, sort_desc = self._parse_sort(query_params, columns)
        if sort_column is None and ('_sort' in query_params or '_sort_desc' in query_params):
            return

        if sort_column is not None:
            col_idx = columns.index(sort_column)
            rows = sorted(rows, key=lambda x: x[col_idx], reverse=sort_desc)

        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(row)

        csv_content = output.getvalue()

        # Send CSV response
        self.send_response(200)
        self.send_header('Content-Type', 'text/csv')
        self.send_header('Content-Disposition', f'attachment; filename="{dataset_id}.csv"')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(csv_content.encode('utf-8'))

    def _parse_filters(self, query_params: dict, columns: list):
        column_set = set(columns)
        valid_comparators = {'exact', 'contains', 'less', 'greater'}
        filters = []

        for raw_key, values in query_params.items():
            if raw_key.startswith('_'):
                continue

            if '__' not in raw_key:
                continue

            if len(values) > 1:
                self.send_error_response(400, f'Duplicate filter key: {raw_key}')
                return None

            parts = raw_key.rsplit('__', 1)
            if len(parts) != 2:
                continue

            column_name, comparator = parts

            if comparator not in valid_comparators:
                self.send_error_response(400, f'Invalid comparator: {comparator}')
                return None

            if column_name not in column_set:
                self.send_error_response(400, f'Unknown column: {column_name}')
                return None

            filter_value = values[0]

            if comparator in ('less', 'greater'):
                try:
                    float(filter_value)
                except (ValueError, TypeError):
                    self.send_error_response(400, f'Non-numeric filter value for {comparator}: {filter_value}')
                    return None

            filters.append((column_name, comparator, filter_value))

        return filters

    def _apply_filters(self, rows: list, columns: list, filters: list) -> list:
        if not filters:
            return rows

        column_index = {col: idx for idx, col in enumerate(columns)}
        result = rows

        for column_name, comparator, filter_value in filters:
            col_idx = column_index[column_name]
            matches = _COMPARATORS[comparator]
            result = [row for row in result if matches(row[col_idx], filter_value)]

        return result

    def handle_dataset(self, dataset_id: str):
        dataset = _datasets.get(dataset_id)
        if dataset is None:
            self.send_error_response(404, 'Dataset not found')
            return

        parsed_url = urlparse(self.path)
        query_params = parse_qs(parsed_url.query, keep_blank_values=True)

        for param in ('_size', '_offset', '_shape', '_sort', '_sort_desc', '_rowid', '_total', '_timeout'):
            if param in query_params and len(query_params[param]) > 1:
                self.send_error_response(400, f'Duplicate parameter: {param}')
                return

        columns = dataset['columns']
        rows = list(dataset['rows'])

        timeout_ms = self._parse_int(query_params, '_timeout', None, min_value=1) \
            if '_timeout' in query_params else None
        if '_timeout' in query_params and timeout_ms is None:
            return

        query_start = time.time()

        def timed_out():
            return timeout_ms is not None and (time.time() - query_start) * 1000 > timeout_ms

        filters = self._parse_filters(query_params, columns)
        if filters is None:
            return

        rows = self._apply_filters(rows, columns, filters)

        if timed_out():
            self.send_error_response(400, 'Query timeout')
            return

        total = len(rows)

        size = self._parse_int(query_params, '_size', 100, min_value=1)
        if size is None:
            return

        offset = self._parse_int(query_params, '_offset', 0, min_value=0)
        if offset is None:
            return

        shape = query_params['_shape'][0] if '_shape' in query_params else 'lists'
        if shape not in ('lists', 'objects'):
            self.send_error_response(400, '_shape must be "lists" or "objects"')
            return

        hide_rowid = self._parse_hide_flag(query_params, '_rowid')
        if hide_rowid is None:
            return

        hide_total = self._parse_hide_flag(query_params, '_total')
        if hide_total is None:
            return

        sort_column, sort_desc = self._parse_sort(query_params, columns)
        if sort_column is None and ('_sort' in query_params or '_sort_desc' in query_params):
            return

        if sort_column is not None:
            col_idx = columns.index(sort_column)
            indexed_rows = list(enumerate(rows))
            indexed_rows.sort(key=lambda x: x[1][col_idx], reverse=sort_desc)
            sort_order = [i for i, _ in indexed_rows]
            rows = [r for _, r in indexed_rows]
        else:
            sort_order = list(range(len(rows)))

        if timed_out():
            self.send_error_response(400, 'Query timeout')
            return

        page_rows = rows[offset:offset + size]
        page_indices = sort_order[offset:offset + size]

        if shape == 'objects':
            result_rows = []
            for i, row in enumerate(page_rows):
                obj = {} if hide_rowid else {'rowid': page_indices[i] + 1}
                for col_idx, col_name in enumerate(columns):
                    obj[col_name] = row[col_idx]
                result_rows.append(obj)
        else:
            result_rows = page_rows

        response = {
            'ok': True,
            'columns': columns,
            'rows': result_rows,
            'query_ms': (time.time() - query_start) * 1000,
        }

        if not hide_total:
            response['total'] = total

        self.send_json_response(200, response)

    def _parse_int(self, query_params: dict, param_name: str,
                   default: int | None, *, min_value: int) -> int | None:
        if param_name not in query_params:
            return default
        try:
            val = int(query_params[param_name][0])
        except ValueError:
            val = min_value - 1
        if val < min_value:
            self.send_error_response(400, f'{param_name} must be an integer >= {min_value}')
            return None
        return val

    def _parse_sort(self, query_params: dict, columns: list):
        column_set = set(columns)
        if '_sort_desc' in query_params:
            col = query_params['_sort_desc'][0]
        elif '_sort' in query_params:
            col = query_params['_sort'][0]
        else:
            return None, False

        if not col or col not in column_set:
            self.send_error_response(400, f'Unknown column: {col}')
            return None, False
        return col, '_sort_desc' in query_params

    def _parse_hide_flag(self, query_params: dict, param_name: str) -> bool | None:
        if param_name not in query_params:
            return False
        if query_params[param_name][0] != 'hide':
            self.send_error_response(400, f'{param_name} must be "hide"')
            return None
        return True

    def log_message(self, format, *args):
        pass


def main():
    parser = argparse.ArgumentParser(description='datagate - CSV data ingestion and query service')
    subparsers = parser.add_subparsers(dest='command')

    start_parser = subparsers.add_parser('start', help='Start the datagate server')
    start_parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    start_parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')

    args = parser.parse_args()

    if args.command == 'start':
        server_address = (args.address, args.port)
        httpd = HTTPServer(server_address, DataGateHandler)
        print(f'datagate running on http://{args.address}:{args.port}')
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print('\nShutting down...')
            httpd.shutdown()
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
