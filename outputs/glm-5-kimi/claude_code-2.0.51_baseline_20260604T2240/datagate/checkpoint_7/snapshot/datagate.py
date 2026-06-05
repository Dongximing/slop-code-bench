#!/usr/bin/env python3
"""
datagate - A CSV to JSON conversion and storage service.
"""

import argparse
import csv
import hashlib
import io
import json
import os
import re
import time
from functools import wraps
from pathlib import Path
from urllib.parse import urlparse

import chardet
import requests
from flask import Flask, jsonify, request, Response
from openpyxl import load_workbook
import xlrd

# ---- Configuration System ----
_TRUE_VALUES = frozenset({'1', 'true', 'yes', 'on'})
_FALSE_VALUES = frozenset({'0', 'false', 'no', 'off'})


def _parse_boolean(value, setting_name):
    """Parse a boolean value string."""
    lowered = value.strip().lower()
    if lowered in _TRUE_VALUES:
        return True
    if lowered in _FALSE_VALUES:
        return False
    print(f"Error: {setting_name} has invalid value '{value}'. "
          f"Acceptable values: {', '.join(sorted(_TRUE_VALUES | _FALSE_VALUES))}",
          flush=True)
    raise SystemExit(1)


def _parse_config_file(filepath):
    """Parse a configuration file with KEY=VALUE lines."""
    config = {}
    try:
        with open(filepath, 'r') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                # Skip blank lines and comments
                if not line or line.startswith('#'):
                    continue
                # Parse KEY=VALUE
                if '=' not in line:
                    print(f"Error: Invalid config line {line_num} in {filepath}: '{line}'",
                          flush=True)
                    raise SystemExit(1)
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()
                if not key:
                    print(f"Error: Empty key in config line {line_num} in {filepath}",
                          flush=True)
                    raise SystemExit(1)
                config[key] = value
    except FileNotFoundError:
        print(f"Error: Config file not found: {filepath}", flush=True)
        raise SystemExit(1)
    except PermissionError:
        print(f"Error: Cannot read config file: {filepath}", flush=True)
        raise SystemExit(1)
    except Exception as e:
        print(f"Error: Failed to read config file {filepath}: {e}", flush=True)
        raise SystemExit(1)
    return config


def _get_config_value(key, default=None):
    """
    Get configuration value following precedence:
    1. Built-in defaults (passed as default parameter)
    2. DATAGATE_CONFIG file
    3. Direct environment variables
    """
    # First check direct environment variable (highest precedence)
    env_value = os.environ.get(key)
    if env_value is not None:
        return env_value

    # Then check config file
    config_path = os.environ.get('DATAGATE_CONFIG')
    if config_path:
        config_data = _parse_config_file(config_path)
        if key in config_data:
            return config_data[key]

    # Finally use built-in default
    return default


def _parse_integer(value, setting_name):
    """Parse an integer value."""
    try:
        return int(value)
    except (ValueError, TypeError):
        print(f"Error: {setting_name} must be an integer, got '{value}'", flush=True)
        raise SystemExit(1)


def _parse_list(value):
    """Parse a comma-separated list value."""
    if not value:
        return []
    return [item.strip() for item in value.split(',') if item.strip()]


class Config:
    """Configuration container with all settings."""

    def __init__(self):
        # MAX_SOURCE_SIZE: integer bytes, unset means no max
        max_size = _get_config_value('MAX_SOURCE_SIZE')
        self.MAX_SOURCE_SIZE = _parse_integer(max_size, 'MAX_SOURCE_SIZE') if max_size is not None else None

        # ORIGIN_ALLOWLIST: list of domain suffixes
        allowlist = _get_config_value('ORIGIN_ALLOWLIST')
        self.ORIGIN_ALLOWLIST = _parse_list(allowlist) if allowlist else None

        # REQUIRE_TLS: boolean, default false
        require_tls = _get_config_value('REQUIRE_TLS', 'false')
        self.REQUIRE_TLS = _parse_boolean(require_tls, 'REQUIRE_TLS')

        # STORAGE_DIR: path, implementation-defined default
        storage_dir = _get_config_value('STORAGE_DIR')
        if storage_dir:
            self.STORAGE_DIR = Path(storage_dir)
        else:
            # Default to 'data' subdirectory
            self.STORAGE_DIR = Path.cwd() / 'data'

        # CACHE_ENABLED: boolean, default true
        cache_enabled = _get_config_value('CACHE_ENABLED', 'true')
        self.CACHE_ENABLED = _parse_boolean(cache_enabled, 'CACHE_ENABLED')

    def is_allowed_origin(self, hostname):
        """Check if hostname matches any allowlist suffix (case-insensitive)."""
        if not self.ORIGIN_ALLOWLIST:
            return True

        hostname_lower = hostname.lower()
        for suffix in self.ORIGIN_ALLOWLIST:
            suffix_lower = suffix.lower()
            # Domain boundary matching: exact match or suffix with domain boundary
            if hostname_lower == suffix_lower:
                return True
            if hostname_lower.endswith('.' + suffix_lower):
                return True
        return False


# Initialize configuration at startup
config = Config()

app = Flask(__name__)

# ---- Storage System ----
# Storage directory path (will be initialized in main())
_storage_dir = None


def _get_storage_path(dataset_id):
    """Get the storage file path for a dataset."""
    if _storage_dir is None:
        return None
    return _storage_dir / f"{dataset_id}.json"


def _save_dataset(dataset_id, dataset):
    """Save a dataset to persistent storage."""
    if _storage_dir is None:
        return
    try:
        filepath = _get_storage_path(dataset_id)
        with open(filepath, 'w') as f:
            json.dump(dataset, f)
    except Exception:
        # Silently ignore storage errors
        pass


def _load_dataset(dataset_id):
    """Load a dataset from persistent storage."""
    if _storage_dir is None:
        return None
    try:
        filepath = _get_storage_path(dataset_id)
        if filepath.exists():
            with open(filepath, 'r') as f:
                return json.load(f)
    except Exception:
        pass
    return None


def _init_storage():
    """Initialize storage directory."""
    global _storage_dir
    _storage_dir = config.STORAGE_DIR
    try:
        _storage_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        # If we can't create storage dir, fall back to memory-only
        _storage_dir = None


def _load_all_datasets():
    """Load all persisted datasets from storage directory."""
    global datasets, _convert_cache
    if _storage_dir is None or not _storage_dir.exists():
        return
    try:
        for filepath in _storage_dir.glob('*.json'):
            try:
                with open(filepath, 'r') as f:
                    dataset = json.load(f)
                dataset_id = filepath.stem
                datasets[dataset_id] = dataset
                # Mark in convert cache
                if config.CACHE_ENABLED:
                    _convert_cache[dataset_id] = True
            except Exception:
                pass
    except Exception:
        pass


# In-memory storage for datasets
datasets = {}
# In-memory storage for cached convert results
# Maps dataset_id (derived from source URL) -> True  (presence-only)
_convert_cache = {}


# ---- Origin Allowlist Middleware ----
@app.before_request
def check_origin_allowlist():
    """Check Referer against ORIGIN_ALLOWLIST before routing."""
    # Skip if no allowlist configured
    if not config.ORIGIN_ALLOWLIST:
        return None

    # Require Referer header
    referer = request.headers.get('Referer')
    if not referer:
        return error_response("Missing Referer header", 403)

    # Parse hostname from Referer
    try:
        parsed = urlparse(referer)
        hostname = parsed.hostname or ''
    except Exception:
        return error_response("Invalid Referer header", 403)

    # Check if hostname matches any allowed suffix
    if not config.is_allowed_origin(hostname):
        return error_response(f"Origin not allowed: {hostname}", 403)

    return None


def make_endpoint_url(dataset_id):
    """Generate endpoint URL based on REQUIRE_TLS setting."""
    if config.REQUIRE_TLS:
        # Absolute HTTPS URL using request host
        host = request.host
        return f"https://{host}/datasets/{dataset_id}"
    else:
        # Relative URL
        return f"/datasets/{dataset_id}"


def make_json_response(data, status=200):
    """Create a JSON response with CORS headers."""
    response = jsonify(data)
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response, status


def error_response(message, status=400):
    """Create an error response."""
    return make_json_response({'ok': False, 'error': message}, status)


def success_response(data):
    """Create a success response."""
    data['ok'] = True
    return make_json_response(data, 200)


def generate_dataset_id(source_url):
    """Generate a deterministic dataset ID from the source URL."""
    return hashlib.sha256(source_url.encode('utf-8')).hexdigest()[:16]


def is_valid_url(url):
    """Check if the URL is valid."""
    try:
        result = urlparse(url)
        return result.scheme in ('http', 'https') and result.netloc
    except Exception:
        return False


def detect_encoding(content_bytes):
    """Detect the encoding of the content."""
    result = chardet.detect(content_bytes)
    return result['encoding'] or 'utf-8'


def detect_delimiter(sample_text):
    """Detect the CSV delimiter from sample text."""
    # Count potential delimiters in the sample
    delimiters = [',', ';', '\t']
    counts = {}

    for delim in delimiters:
        # Count occurrences in each line
        lines = sample_text.split('\n')[:10]  # Check first 10 lines
        line_counts = [line.count(delim) for line in lines if line.strip()]
        if line_counts:
            # A good delimiter should have consistent counts across lines
            if len(set(line_counts)) == 1 and line_counts[0] > 0:
                counts[delim] = line_counts[0]
            elif line_counts[0] > 0:
                # Allow some variation but require majority
                max_count = max(line_counts)
                min_count = min(line_counts)
                if max_count > 0 and min_count > 0:
                    counts[delim] = sum(line_counts)

    if not counts:
        # Fallback: try comma as default
        return ','

    # Return the delimiter with highest count
    return max(counts, key=counts.get)


def is_time_like(value):
    """Check if a value looks like a time (e.g., '08:30', '9:15', '12:00')."""
    if not isinstance(value, str):
        return False
    # Match time patterns like HH:MM, H:MM, HH:MM:SS, H:MM:SS
    time_pattern = r'^\d{1,2}:\d{2}(:\d{2})?$'
    return bool(re.match(time_pattern, value.strip()))


def compute_enrichment_metadata(columns, rows, filetype):
    """
    Compute enrichment metadata for a dataset.

    For CSV datasets:
    - dataset_summary with filetype, row_count, column_count
    - column_details keyed by column name with type, distinct_count, missing_count

    For spreadsheet (excel) datasets:
    - dataset_summary with filetype: "excel"
    """
    # Build dataset_summary
    dataset_summary = {
        'filetype': filetype,
        'row_count': len(rows),
        'column_count': len(columns)
    }

    result = {
        'dataset_summary': dataset_summary
    }

    # For CSV, also include column_details
    if filetype == 'csv':
        column_details = {}
        for col_idx, col_name in enumerate(columns):
            # Extract all values for this column
            col_values = []
            missing_count = 0
            for row in rows:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val == '' or val is None:
                        missing_count += 1
                    else:
                        col_values.append(val)
                else:
                    missing_count += 1

            # Determine type: text, number, integer, float
            col_type = 'text'
            if col_values:
                # Check if all non-missing values are numbers
                all_int = True
                all_float = True
                has_numeric = False

                for val in col_values:
                    if isinstance(val, int):
                        has_numeric = True
                        # integer stays integer
                    elif isinstance(val, float):
                        has_numeric = True
                        all_int = False
                    elif isinstance(val, str):
                        try:
                            # Try integer first
                            int(val)
                            has_numeric = True
                            # It's an integer string
                        except (ValueError, TypeError):
                            try:
                                float(val)
                                has_numeric = True
                                all_int = False
                            except (ValueError, TypeError):
                                all_int = False
                                all_float = False
                                break
                    else:
                        all_int = False
                        all_float = False
                        break

                if has_numeric and (all_int or all_float):
                    if all_int:
                        col_type = 'integer'
                    else:
                        col_type = 'float'

            # Count distinct values
            distinct_values = set()
            for val in col_values:
                # Use string representation for hashing to handle different types
                distinct_values.add(str(val))
            distinct_count = len(distinct_values)

            column_details[col_name] = {
                'type': col_type,
                'distinct_count': distinct_count,
                'missing_count': missing_count
            }

        result['column_details'] = column_details

    return result


def infer_type(value):
    """Infer the JSON type for a value."""
    if value == '' or value is None:
        return ''

    # Keep time-like values as strings
    if is_time_like(value):
        return value

    # Try integer
    try:
        int_val = int(value)
        # Check if it's truly an integer (not float notation)
        if '.' not in str(value) and 'e' not in str(value).lower():
            return int_val
    except (ValueError, TypeError):
        pass

    # Try float
    try:
        float_val = float(value)
        return float_val
    except (ValueError, TypeError):
        pass

    # Return as string
    return value


def parse_csv(content_bytes, charset=None):
    """Parse CSV content and return columns and rows."""
    # Decode content
    if charset:
        try:
            content = content_bytes.decode(charset)
        except (UnicodeDecodeError, LookupError, AttributeError) as e:
            raise ValueError(f"Invalid charset '{charset}': {str(e)}")
    else:
        encoding = detect_encoding(content_bytes)
        try:
            content = content_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            # Fallback to utf-8 with error handling
            content = content_bytes.decode('utf-8', errors='replace')

    if not content.strip():
        raise ValueError("Empty content")

    # Detect delimiter
    sample = content[:min(len(content), 8192)]
    delimiter = detect_delimiter(sample)

    # Parse CSV
    try:
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = list(reader)
    except Exception as e:
        raise ValueError(f"Failed to parse CSV: {str(e)}")

    if len(rows) < 2:
        raise ValueError("CSV must have at least one header row and one data row")

    # Extract columns from first row
    columns = rows[0]

    if not columns or all(col.strip() == '' for col in columns):
        raise ValueError("CSV has no valid headers")

    # Process data rows
    data_rows = []
    for row in rows[1:]:
        # Infer types for each value
        typed_row = []
        for i, value in enumerate(row):
            if i < len(columns):
                typed_row.append(infer_type(value))
        # Ensure row has same number of columns as header
        while len(typed_row) < len(columns):
            typed_row.append('')
        data_rows.append(typed_row[:len(columns)])

    return columns, data_rows


def parse_xlsx(content_bytes):
    """Parse XLSX content and return columns and rows."""
    try:
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([cell if cell is not None else '' for cell in row])

        wb.close()

        if len(rows) < 2:
            raise ValueError("Spreadsheet must have at least one header row and one data row")

        columns = [str(col) if col is not None else '' for col in rows[0]]

        if not columns or all(col.strip() == '' for col in columns):
            raise ValueError("Spreadsheet has no valid headers")

        # Process data rows
        data_rows = []
        for row in rows[1:]:
            typed_row = []
            for i, value in enumerate(row):
                if i < len(columns):
                    typed_row.append(infer_type(str(value) if value is not None else ''))
            while len(typed_row) < len(columns):
                typed_row.append('')
            data_rows.append(typed_row[:len(columns)])

        return columns, data_rows
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Failed to parse XLSX: {str(e)}")


def parse_xls(content_bytes):
    """Parse XLS content and return columns and rows."""
    try:
        wb = xlrd.open_workbook(file_contents=content_bytes)
        ws = wb.sheet_by_index(0)

        if ws.nrows < 2:
            raise ValueError("Spreadsheet must have at least one header row and one data row")

        columns = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]

        if not columns or all(col.strip() == '' for col in columns):
            raise ValueError("Spreadsheet has no valid headers")

        data_rows = []
        for row_idx in range(1, ws.nrows):
            typed_row = []
            for col_idx in range(len(columns)):
                cell_value = ws.cell_value(row_idx, col_idx)
                typed_row.append(infer_type(str(cell_value) if cell_value is not None else ''))
            while len(typed_row) < len(columns):
                typed_row.append('')
            data_rows.append(typed_row[:len(columns)])

        return columns, data_rows
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Failed to parse XLS: {str(e)}")


def parse_content(content_bytes, charset=None, filename=None):
    """Parse content based on format (CSV, XLS, XLSX) and return columns, rows, and filetype."""
    # Try to detect format from filename or content
    if filename:
        filename_lower = filename.lower()
        if filename_lower.endswith('.xlsx'):
            columns, rows = parse_xlsx(content_bytes)
            return columns, rows, 'excel'
        elif filename_lower.endswith('.xls'):
            columns, rows = parse_xls(content_bytes)
            return columns, rows, 'excel'

    # Try to parse as XLSX first (binary format)
    try:
        columns, rows = parse_xlsx(content_bytes)
        return columns, rows, 'excel'
    except:
        pass

    # Try to parse as XLS
    try:
        columns, rows = parse_xls(content_bytes)
        return columns, rows, 'excel'
    except:
        pass

    # Fall back to CSV
    columns, rows = parse_csv(content_bytes, charset)
    return columns, rows, 'csv'


@app.route('/convert', methods=['GET', 'OPTIONS'])
def convert():
    """Convert a remote CSV/spreadsheet file to a dataset."""
    if request.method == 'OPTIONS':
        return make_json_response({'ok': True}, 200)

    source = request.args.get('source')
    charset = request.args.get('charset')

    # ---- force parameter validation (presence flag) ----
    force_param_count = len(request.args.getlist('force'))
    if force_param_count > 1:
        return error_response("Duplicate 'force' parameter", 400)
    force = force_param_count == 1

    # ---- enrich parameter validation ----
    enrich_param_count = len(request.args.getlist('enrich'))
    if enrich_param_count > 1:
        return error_response("Duplicate 'enrich' parameter", 400)
    enrich = request.args.get('enrich') == 'yes'

    # Validate source parameter
    if not source:
        return error_response("Missing 'source' parameter", 400)

    if not is_valid_url(source):
        return error_response("Invalid URL", 400)

    # Validate charset if provided
    if charset:
        try:
            # Test if charset is valid
            ''.encode(charset)
        except LookupError:
            return error_response(f"Unsupported charset: {charset}", 400)

    # Generate dataset ID (deterministic from source URL)
    dataset_id = generate_dataset_id(source)

    # ---- Caching logic with enrichment state handling ----
    existing_dataset = datasets.get(dataset_id)
    is_enriched = existing_dataset is not None and 'enriched' in existing_dataset

    if config.CACHE_ENABLED and not force and dataset_id in _convert_cache:
        # Cache hit - check enrichment state
        if enrich:
            # enrich=yes on cached dataset
            if is_enriched:
                # Already enriched - can use cache
                return success_response({'endpoint': make_endpoint_url(dataset_id)})
            else:
                # Not enriched - need to re-ingest with enrichment
                pass  # Fall through to re-ingestion
        else:
            # No enrichment requested - use cache as-is
            return success_response({'endpoint': make_endpoint_url(dataset_id)})

    # Cache miss, forced, or enrichment upgrade needed: need to (re-)ingest
    # Fetch the file
    try:
        response = requests.get(source, timeout=30)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        # Re-ingestion failure: keep existing error codes and envelope
        error_msg = f"Source unreachable: {str(e)}"
        status = 404
        # If forced re-ingestion fails and a prior dataset exists, keep it queryable
        # (dataset already exists in datasets dict, nothing to do)
        return error_response(error_msg, status)

    # ---- MAX_SOURCE_SIZE enforcement ----
    if config.MAX_SOURCE_SIZE is not None:
        content_length = len(response.content)
        if content_length > config.MAX_SOURCE_SIZE:
            return error_response(f"File size {content_length} exceeds maximum {config.MAX_SOURCE_SIZE}", 400)

    # Check content type hints
    content_type = response.headers.get('Content-Type', '')
    if 'json' in content_type.lower() or 'xml' in content_type.lower() or 'html' in content_type.lower():
        # Re-ingestion failure: keep existing error codes and envelope
        # If forced re-ingestion fails and a prior dataset exists, keep it queryable
        return error_response("Non-tabular content", 400)

    # Extract filename from URL
    filename = source.split('/')[-1] if '/' in source else source

    # Parse the content
    try:
        columns, rows, filetype = parse_content(response.content, charset, filename)
    except ValueError as e:
        # Re-ingestion failure: keep existing error codes and envelope
        # If forced re-ingestion fails and a prior dataset exists, keep it queryable
        return error_response(str(e), 400)
    except Exception as e:
        # Re-ingestion failure: keep existing error codes and envelope
        # If forced re-ingestion fails and a prior dataset exists, keep it queryable
        return error_response(f"Failed to parse file: {str(e)}", 400)

    # Store the dataset
    dataset = {
        'columns': columns,
        'rows': rows,
        'source': source
    }

    # Add enrichment metadata if requested
    if enrich:
        try:
            enrichment = compute_enrichment_metadata(columns, rows, filetype)
            dataset['enriched'] = enrichment
        except Exception as e:
            # Enrichment failure - return standard JSON error, do not downgrade
            return error_response(f"Enrichment failed: {str(e)}", 400)

    datasets[dataset_id] = dataset

    # Persist to storage
    _save_dataset(dataset_id, dataset)

    # Mark in cache (only if caching enabled)
    if config.CACHE_ENABLED:
        _convert_cache[dataset_id] = True

    return success_response({'endpoint': make_endpoint_url(dataset_id)})


@app.route('/datasets/<dataset_id>', methods=['GET', 'OPTIONS'])
def get_dataset(dataset_id):
    """Get a stored dataset."""
    if request.method == 'OPTIONS':
        return make_json_response({'ok': True}, 200)

    start_time = time.time()

    if dataset_id not in datasets:
        return error_response("Dataset not found", 404)

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = dataset['rows']

    # Check for duplicate control parameters
    control_params = ['_size', '_offset', '_shape', '_sort', '_sort_desc', '_rowid', '_total']
    for param in control_params:
        param_list = request.args.getlist(param)
        if len(param_list) > 1:
            return error_response(f"Duplicate parameter: {param}", 400)

    # Parse and validate filters
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filters = {}  # Maps (column, comparator) tuple to value
    seen_filter_keys = set()

    for key in request.args.keys():
        # Skip control parameters (starting with _)
        if key.startswith('_'):
            continue

        # Check if this is a filter parameter (must contain __)
        if '__' not in key:
            continue

        # Check for duplicate filter keys
        param_list = request.args.getlist(key)
        if len(param_list) > 1:
            return error_response(f"Duplicate filter key: {key}", 400)

        value = param_list[0]

        # Split on __ to get column and comparator
        parts = key.split('__', 1)
        column = parts[0]
        comparator = parts[1]

        # Check for empty comparator (trailing __ or just __)
        if comparator == '':
            return error_response(f"Invalid comparator: {comparator}", 400)

        # Validate comparator
        if comparator not in valid_comparators:
            return error_response(f"Invalid comparator: {comparator}", 400)

        # Validate column exists
        if column not in columns:
            return error_response(f"Unknown filter column: {column}", 400)

        # Track filter key to prevent duplicates
        filter_key = (column, comparator)
        seen_filter_keys.add(filter_key)

        # For less/greater comparators, validate that filter value is numeric
        if comparator in ('less', 'greater'):
            try:
                float(value)
            except (ValueError, TypeError):
                return error_response(f"Comparator target not numeric: {value}", 400)

        filters[filter_key] = value

    # Parse pagination parameters
    size = request.args.get('_size', '100')
    offset = request.args.get('_offset', '0')

    # Validate _size
    try:
        size = int(size)
        if size <= 0:
            return error_response("_size must be a positive integer", 400)
    except (ValueError, TypeError):
        return error_response("_size must be a positive integer", 400)

    # Validate _offset
    try:
        offset = int(offset)
        if offset < 0:
            return error_response("_offset must be a non-negative integer", 400)
    except (ValueError, TypeError):
        return error_response("_offset must be a non-negative integer", 400)

    # Parse shape parameter
    shape = request.args.get('_shape', 'lists')
    if shape not in ('lists', 'objects'):
        return error_response("_shape must be 'lists' or 'objects'", 400)

    # Parse sorting parameters
    sort = request.args.get('_sort')
    sort_desc = request.args.get('_sort_desc')

    # Validate sort parameters
    if sort is not None:
        if sort == '' or sort not in columns:
            return error_response(f"Unknown column for sorting: {sort}", 400)

    if sort_desc is not None:
        if sort_desc == '' or sort_desc not in columns:
            return error_response(f"Unknown column for sorting: {sort_desc}", 400)

    # Parse visibility toggles
    rowid_toggle = request.args.get('_rowid')
    total_toggle = request.args.get('_total')

    if rowid_toggle is not None and rowid_toggle != 'hide':
        return error_response("_rowid must be 'hide'", 400)

    if total_toggle is not None and total_toggle != 'hide':
        return error_response("_total must be 'hide'", 400)

    # Apply filtering before sorting
    if filters:
        filtered_rows = []
        for row in rows:
            match = True
            for (column, comparator), filter_value in filters.items():
                col_index = columns.index(column)
                cell_value = row[col_index] if col_index < len(row) else ''

                if comparator == 'exact':
                    # Case-sensitive string equality
                    if str(cell_value) != filter_value:
                        match = False
                        break
                elif comparator == 'contains':
                    # Case-sensitive substring
                    if filter_value not in str(cell_value):
                        match = False
                        break
                elif comparator == 'less':
                    # Numeric strict less
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num < filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        # Non-numeric stored value - row not matched
                        match = False
                        break
                elif comparator == 'greater':
                    # Numeric strict greater
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num > filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        # Non-numeric stored value - row not matched
                        match = False
                        break

            if match:
                filtered_rows.append(row)
        rows = filtered_rows

    # Calculate total before pagination (after filtering)
    total = len(rows)

    # Apply sorting (stable sort) - convert values to strings for comparison to handle mixed types
    if sort_desc is not None:
        # Sort descending (takes precedence over _sort)
        col_index = columns.index(sort_desc)
        # Use enumerate to maintain stability
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''), reverse=True)
        rows = [r[1] for r in indexed_rows]
    elif sort is not None:
        # Sort ascending
        col_index = columns.index(sort)
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''))
        rows = [r[1] for r in indexed_rows]

    # Apply pagination
    if offset >= len(rows):
        rows = []
    else:
        rows = rows[offset:offset + size]

    # Build response
    response_data = {
        'columns': columns,
        'query_ms': round((time.time() - start_time) * 1000, 1)
    }

    # Add total unless hidden
    if total_toggle != 'hide':
        response_data['total'] = total

    # Add enrichment metadata if present (non-enriched datasets omit these fields)
    if 'enriched' in dataset:
        response_data['dataset_summary'] = dataset['enriched']['dataset_summary']
        if 'column_details' in dataset['enriched']:
            response_data['column_details'] = dataset['enriched']['column_details']

    # Format rows based on shape
    if shape == 'objects':
        formatted_rows = []
        for i, row in enumerate(rows):
            obj = {'rowid': offset + i + 1}  # 1-based rowid
            for j, col in enumerate(columns):
                obj[col] = row[j] if j < len(row) else ''
            formatted_rows.append(obj)

        if rowid_toggle == 'hide':
            for obj in formatted_rows:
                del obj['rowid']

        response_data['rows'] = formatted_rows
    else:
        # lists shape (default)
        response_data['rows'] = rows

    return success_response(response_data)


@app.route('/upload', methods=['POST', 'OPTIONS'])
def upload():
    """Upload a file (CSV, XLS, XLSX) and create a dataset."""
    if request.method == 'OPTIONS':
        return make_json_response({'ok': True}, 200)

    # Check if request is multipart
    if not request.files:
        return error_response("Non-multipart upload", 415)

    # Try to get file from 'file' or 'attachment' field
    uploaded_file = request.files.get('file') or request.files.get('attachment')

    if not uploaded_file:
        return error_response("Missing file field", 400)

    # Read file content
    content = uploaded_file.read()
    filename = uploaded_file.filename or ''

    # ---- MAX_SOURCE_SIZE enforcement ----
    if config.MAX_SOURCE_SIZE is not None:
        content_length = len(content)
        if content_length > config.MAX_SOURCE_SIZE:
            return error_response(f"File size {content_length} exceeds maximum {config.MAX_SOURCE_SIZE}", 400)

    # Generate dataset ID from file content hash (deterministic)
    dataset_id = hashlib.sha256(content).hexdigest()[:16]

    # Parse the content
    charset = request.form.get('charset')

    try:
        columns, rows, filetype = parse_content(content, charset, filename)
    except ValueError as e:
        return error_response(str(e), 400)
    except Exception as e:
        return error_response(f"Failed to parse file: {str(e)}", 400)

    # Store the dataset
    dataset = {
        'columns': columns,
        'rows': rows,
        'source': f'upload://{filename}'
    }
    datasets[dataset_id] = dataset

    # Persist to storage
    _save_dataset(dataset_id, dataset)

    return success_response({'endpoint': make_endpoint_url(dataset_id)})


@app.route('/datasets/<dataset_id>/export', methods=['GET', 'OPTIONS'])
def export_dataset(dataset_id):
    """Export a dataset as CSV."""
    if request.method == 'OPTIONS':
        response = Response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200

    if dataset_id not in datasets:
        return error_response("Dataset not found", 404)

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = dataset['rows']

    # Parse query parameters for filtering, sorting, and pagination
    # (same as /datasets/<id> but only for row selection)

    # Parse and validate filters
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filters = {}

    for key in request.args.keys():
        # Skip control parameters (starting with _)
        if key.startswith('_'):
            continue

        # Check if this is a filter parameter (must contain __)
        if '__' not in key:
            continue

        value = request.args.get(key)
        if value is None:
            continue

        # Split on __ to get column and comparator
        parts = key.split('__', 1)
        column = parts[0]
        comparator = parts[1]

        # Validate comparator
        if comparator not in valid_comparators:
            continue

        # Validate column exists
        if column not in columns:
            continue

        filters[(column, comparator)] = value

    # Parse pagination parameters for export
    size = request.args.get('_size')
    offset = request.args.get('_offset', '0')

    if size is not None:
        try:
            size = int(size)
            if size <= 0:
                return error_response("_size must be a positive integer", 400)
        except (ValueError, TypeError):
            return error_response("_size must be a positive integer", 400)
    else:
        size = None  # No limit

    try:
        offset = int(offset)
        if offset < 0:
            return error_response("_offset must be a non-negative integer", 400)
    except (ValueError, TypeError):
        return error_response("_offset must be a non-negative integer", 400)

    # Parse sorting parameters
    sort = request.args.get('_sort')
    sort_desc = request.args.get('_sort_desc')

    if sort is not None:
        if sort == '' or sort not in columns:
            sort = None

    if sort_desc is not None:
        if sort_desc == '' or sort_desc not in columns:
            sort_desc = None

    # Apply filtering before sorting
    if filters:
        filtered_rows = []
        for row in rows:
            match = True
            for (column, comparator), filter_value in filters.items():
                col_index = columns.index(column)
                cell_value = row[col_index] if col_index < len(row) else ''

                if comparator == 'exact':
                    if str(cell_value) != filter_value:
                        match = False
                        break
                elif comparator == 'contains':
                    if filter_value not in str(cell_value):
                        match = False
                        break
                elif comparator == 'less':
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num < filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        match = False
                        break
                elif comparator == 'greater':
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num > filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        match = False
                        break

            if match:
                filtered_rows.append(row)
        rows = filtered_rows

    # Apply sorting (stable sort)
    if sort_desc is not None:
        col_index = columns.index(sort_desc)
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''), reverse=True)
        rows = [r[1] for r in indexed_rows]
    elif sort is not None:
        col_index = columns.index(sort)
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''))
        rows = [r[1] for r in indexed_rows]

    # Apply pagination
    if offset >= len(rows):
        rows = []
    elif size is not None:
        rows = rows[offset:offset + size]
    else:
        rows = rows[offset:]

    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)

    csv_content = output.getvalue()

    # Create response with appropriate headers
    response = Response(
        csv_content,
        mimetype='text/csv',
        headers={
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename="{dataset_id}.csv"',
            'Access-Control-Allow-Origin': '*'
        }
    )

    return response, 200


@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors."""
    return error_response("Not found", 404)


@app.errorhandler(500)
def internal_error(e):
    """Handle 500 errors."""
    return error_response("Internal server error", 500)


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description='datagate - CSV to JSON conversion service')
    subparsers = parser.add_subparsers(dest='command')

    start_parser = subparsers.add_parser('start', help='Start the server')
    start_parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    start_parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')

    args = parser.parse_args()

    if args.command == 'start':
        # Initialize storage directory
        _init_storage()
        # Load persisted datasets
        _load_all_datasets()

        print(f"Starting datagate on {args.address}:{args.port}")
        app.run(host=args.address, port=args.port)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
