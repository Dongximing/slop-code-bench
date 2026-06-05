#!/usr/bin/env python3
"""
datagate - A CSV to JSON conversion and storage service.
"""

import argparse
import csv
import hashlib
import io
import json
import re
import time
from functools import wraps
from urllib.parse import urlparse

import chardet
import requests
from flask import Flask, jsonify, request, Response
from openpyxl import load_workbook
import xlrd

app = Flask(__name__)

# In-memory storage for datasets
datasets = {}


def make_json_response(data, status=200):
    """Create a JSON response with CORS headers."""
    response = jsonify(data)
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response, status


def error_response(message, status=400):
    """Create an error response."""
    return make_json_response({'ok': False, 'error': message}, status)


def success_response(data):
    """Create a success response."""
    data['ok'] = True
    return make_json_response(data, 200)


def generate_dataset_id(source_url):
    """Generate a deterministic dataset ID from the source URL."""
    return hashlib.sha256(source_url.encode('utf-8')).hexdigest()[:16]


def is_valid_url(url):
    """Check if the URL is valid."""
    try:
        result = urlparse(url)
        return result.scheme in ('http', 'https') and result.netloc
    except Exception:
        return False


def detect_encoding(content_bytes):
    """Detect the encoding of the content."""
    result = chardet.detect(content_bytes)
    return result['encoding'] or 'utf-8'


def detect_delimiter(sample_text):
    """Detect the CSV delimiter from sample text."""
    # Count potential delimiters in the sample
    delimiters = [',', ';', '\t']
    counts = {}

    for delim in delimiters:
        # Count occurrences in each line
        lines = sample_text.split('\n')[:10]  # Check first 10 lines
        line_counts = [line.count(delim) for line in lines if line.strip()]
        if line_counts:
            # A good delimiter should have consistent counts across lines
            if len(set(line_counts)) == 1 and line_counts[0] > 0:
                counts[delim] = line_counts[0]
            elif line_counts[0] > 0:
                # Allow some variation but require majority
                max_count = max(line_counts)
                min_count = min(line_counts)
                if max_count > 0 and min_count > 0:
                    counts[delim] = sum(line_counts)

    if not counts:
        # Fallback: try comma as default
        return ','

    # Return the delimiter with highest count
    return max(counts, key=counts.get)


def is_time_like(value):
    """Check if a value looks like a time (e.g., '08:30', '9:15', '12:00')."""
    if not isinstance(value, str):
        return False
    # Match time patterns like HH:MM, H:MM, HH:MM:SS, H:MM:SS
    time_pattern = r'^\d{1,2}:\d{2}(:\d{2})?$'
    return bool(re.match(time_pattern, value.strip()))


def infer_type(value):
    """Infer the JSON type for a value."""
    if value == '' or value is None:
        return ''

    # Keep time-like values as strings
    if is_time_like(value):
        return value

    # Try integer
    try:
        int_val = int(value)
        # Check if it's truly an integer (not float notation)
        if '.' not in str(value) and 'e' not in str(value).lower():
            return int_val
    except (ValueError, TypeError):
        pass

    # Try float
    try:
        float_val = float(value)
        return float_val
    except (ValueError, TypeError):
        pass

    # Return as string
    return value


def parse_csv(content_bytes, charset=None):
    """Parse CSV content and return columns and rows."""
    # Decode content
    if charset:
        try:
            content = content_bytes.decode(charset)
        except (UnicodeDecodeError, LookupError, AttributeError) as e:
            raise ValueError(f"Invalid charset '{charset}': {str(e)}")
    else:
        encoding = detect_encoding(content_bytes)
        try:
            content = content_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            # Fallback to utf-8 with error handling
            content = content_bytes.decode('utf-8', errors='replace')

    if not content.strip():
        raise ValueError("Empty content")

    # Detect delimiter
    sample = content[:min(len(content), 8192)]
    delimiter = detect_delimiter(sample)

    # Parse CSV
    try:
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = list(reader)
    except Exception as e:
        raise ValueError(f"Failed to parse CSV: {str(e)}")

    if len(rows) < 2:
        raise ValueError("CSV must have at least one header row and one data row")

    # Extract columns from first row
    columns = rows[0]

    if not columns or all(col.strip() == '' for col in columns):
        raise ValueError("CSV has no valid headers")

    # Process data rows
    data_rows = []
    for row in rows[1:]:
        # Infer types for each value
        typed_row = []
        for i, value in enumerate(row):
            if i < len(columns):
                typed_row.append(infer_type(value))
        # Ensure row has same number of columns as header
        while len(typed_row) < len(columns):
            typed_row.append('')
        data_rows.append(typed_row[:len(columns)])

    return columns, data_rows


def parse_xlsx(content_bytes):
    """Parse XLSX content and return columns and rows."""
    try:
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([cell if cell is not None else '' for cell in row])

        wb.close()

        if len(rows) < 2:
            raise ValueError("Spreadsheet must have at least one header row and one data row")

        columns = [str(col) if col is not None else '' for col in rows[0]]

        if not columns or all(col.strip() == '' for col in columns):
            raise ValueError("Spreadsheet has no valid headers")

        # Process data rows
        data_rows = []
        for row in rows[1:]:
            typed_row = []
            for i, value in enumerate(row):
                if i < len(columns):
                    typed_row.append(infer_type(str(value) if value is not None else ''))
            while len(typed_row) < len(columns):
                typed_row.append('')
            data_rows.append(typed_row[:len(columns)])

        return columns, data_rows
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Failed to parse XLSX: {str(e)}")


def parse_xls(content_bytes):
    """Parse XLS content and return columns and rows."""
    try:
        wb = xlrd.open_workbook(file_contents=content_bytes)
        ws = wb.sheet_by_index(0)

        if ws.nrows < 2:
            raise ValueError("Spreadsheet must have at least one header row and one data row")

        columns = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]

        if not columns or all(col.strip() == '' for col in columns):
            raise ValueError("Spreadsheet has no valid headers")

        data_rows = []
        for row_idx in range(1, ws.nrows):
            typed_row = []
            for col_idx in range(len(columns)):
                cell_value = ws.cell_value(row_idx, col_idx)
                typed_row.append(infer_type(str(cell_value) if cell_value is not None else ''))
            while len(typed_row) < len(columns):
                typed_row.append('')
            data_rows.append(typed_row[:len(columns)])

        return columns, data_rows
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Failed to parse XLS: {str(e)}")


def parse_content(content_bytes, charset=None, filename=None):
    """Parse content based on format (CSV, XLS, XLSX) and return columns and rows."""
    # Try to detect format from filename or content
    if filename:
        filename_lower = filename.lower()
        if filename_lower.endswith('.xlsx'):
            return parse_xlsx(content_bytes)
        elif filename_lower.endswith('.xls'):
            return parse_xls(content_bytes)

    # Try to parse as XLSX first (binary format)
    try:
        return parse_xlsx(content_bytes)
    except:
        pass

    # Try to parse as XLS
    try:
        return parse_xls(content_bytes)
    except:
        pass

    # Fall back to CSV
    return parse_csv(content_bytes, charset)


@app.route('/convert', methods=['GET', 'OPTIONS'])
def convert():
    """Convert a remote CSV/spreadsheet file to a dataset."""
    if request.method == 'OPTIONS':
        return make_json_response({'ok': True}, 200)

    source = request.args.get('source')
    charset = request.args.get('charset')

    # Validate source parameter
    if not source:
        return error_response("Missing 'source' parameter", 400)

    if not is_valid_url(source):
        return error_response("Invalid URL", 400)

    # Validate charset if provided
    if charset:
        try:
            # Test if charset is valid
            ''.encode(charset)
        except LookupError:
            return error_response(f"Unsupported charset: {charset}", 400)

    # Fetch the file
    try:
        response = requests.get(source, timeout=30)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        return error_response(f"Source unreachable: {str(e)}", 404)

    # Check content type hints
    content_type = response.headers.get('Content-Type', '')
    if 'json' in content_type.lower() or 'xml' in content_type.lower() or 'html' in content_type.lower():
        return error_response("Non-tabular content", 400)

    # Extract filename from URL
    filename = source.split('/')[-1] if '/' in source else source

    # Parse the content
    try:
        columns, rows = parse_content(response.content, charset, filename)
    except ValueError as e:
        return error_response(str(e), 400)
    except Exception as e:
        return error_response(f"Failed to parse file: {str(e)}", 400)

    # Generate dataset ID
    dataset_id = generate_dataset_id(source)

    # Store the dataset
    datasets[dataset_id] = {
        'columns': columns,
        'rows': rows,
        'source': source
    }

    return success_response({'endpoint': f'/datasets/{dataset_id}'})


@app.route('/datasets/<dataset_id>', methods=['GET', 'OPTIONS'])
def get_dataset(dataset_id):
    """Get a stored dataset."""
    if request.method == 'OPTIONS':
        return make_json_response({'ok': True}, 200)

    start_time = time.time()

    if dataset_id not in datasets:
        return error_response("Dataset not found", 404)

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = dataset['rows']

    # Check for duplicate control parameters
    control_params = ['_size', '_offset', '_shape', '_sort', '_sort_desc', '_rowid', '_total']
    for param in control_params:
        param_list = request.args.getlist(param)
        if len(param_list) > 1:
            return error_response(f"Duplicate parameter: {param}", 400)

    # Parse and validate filters
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filters = {}  # Maps (column, comparator) tuple to value
    seen_filter_keys = set()

    for key in request.args.keys():
        # Skip control parameters (starting with _)
        if key.startswith('_'):
            continue

        # Check if this is a filter parameter (must contain __)
        if '__' not in key:
            continue

        # Check for duplicate filter keys
        param_list = request.args.getlist(key)
        if len(param_list) > 1:
            return error_response(f"Duplicate filter key: {key}", 400)

        value = param_list[0]

        # Split on __ to get column and comparator
        parts = key.split('__', 1)
        column = parts[0]
        comparator = parts[1]

        # Check for empty comparator (trailing __ or just __)
        if comparator == '':
            return error_response(f"Invalid comparator: {comparator}", 400)

        # Validate comparator
        if comparator not in valid_comparators:
            return error_response(f"Invalid comparator: {comparator}", 400)

        # Validate column exists
        if column not in columns:
            return error_response(f"Unknown filter column: {column}", 400)

        # Track filter key to prevent duplicates
        filter_key = (column, comparator)
        seen_filter_keys.add(filter_key)

        # For less/greater comparators, validate that filter value is numeric
        if comparator in ('less', 'greater'):
            try:
                float(value)
            except (ValueError, TypeError):
                return error_response(f"Comparator target not numeric: {value}", 400)

        filters[filter_key] = value

    # Parse pagination parameters
    size = request.args.get('_size', '100')
    offset = request.args.get('_offset', '0')

    # Validate _size
    try:
        size = int(size)
        if size <= 0:
            return error_response("_size must be a positive integer", 400)
    except (ValueError, TypeError):
        return error_response("_size must be a positive integer", 400)

    # Validate _offset
    try:
        offset = int(offset)
        if offset < 0:
            return error_response("_offset must be a non-negative integer", 400)
    except (ValueError, TypeError):
        return error_response("_offset must be a non-negative integer", 400)

    # Parse shape parameter
    shape = request.args.get('_shape', 'lists')
    if shape not in ('lists', 'objects'):
        return error_response("_shape must be 'lists' or 'objects'", 400)

    # Parse sorting parameters
    sort = request.args.get('_sort')
    sort_desc = request.args.get('_sort_desc')

    # Validate sort parameters
    if sort is not None:
        if sort == '' or sort not in columns:
            return error_response(f"Unknown column for sorting: {sort}", 400)

    if sort_desc is not None:
        if sort_desc == '' or sort_desc not in columns:
            return error_response(f"Unknown column for sorting: {sort_desc}", 400)

    # Parse visibility toggles
    rowid_toggle = request.args.get('_rowid')
    total_toggle = request.args.get('_total')

    if rowid_toggle is not None and rowid_toggle != 'hide':
        return error_response("_rowid must be 'hide'", 400)

    if total_toggle is not None and total_toggle != 'hide':
        return error_response("_total must be 'hide'", 400)

    # Apply filtering before sorting
    if filters:
        filtered_rows = []
        for row in rows:
            match = True
            for (column, comparator), filter_value in filters.items():
                col_index = columns.index(column)
                cell_value = row[col_index] if col_index < len(row) else ''

                if comparator == 'exact':
                    # Case-sensitive string equality
                    if str(cell_value) != filter_value:
                        match = False
                        break
                elif comparator == 'contains':
                    # Case-sensitive substring
                    if filter_value not in str(cell_value):
                        match = False
                        break
                elif comparator == 'less':
                    # Numeric strict less
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num < filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        # Non-numeric stored value - row not matched
                        match = False
                        break
                elif comparator == 'greater':
                    # Numeric strict greater
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num > filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        # Non-numeric stored value - row not matched
                        match = False
                        break

            if match:
                filtered_rows.append(row)
        rows = filtered_rows

    # Calculate total before pagination (after filtering)
    total = len(rows)

    # Apply sorting (stable sort) - convert values to strings for comparison to handle mixed types
    if sort_desc is not None:
        # Sort descending (takes precedence over _sort)
        col_index = columns.index(sort_desc)
        # Use enumerate to maintain stability
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''), reverse=True)
        rows = [r[1] for r in indexed_rows]
    elif sort is not None:
        # Sort ascending
        col_index = columns.index(sort)
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''))
        rows = [r[1] for r in indexed_rows]

    # Apply pagination
    if offset >= len(rows):
        rows = []
    else:
        rows = rows[offset:offset + size]

    # Build response
    response_data = {
        'columns': columns,
        'query_ms': round((time.time() - start_time) * 1000, 1)
    }

    # Add total unless hidden
    if total_toggle != 'hide':
        response_data['total'] = total

    # Format rows based on shape
    if shape == 'objects':
        formatted_rows = []
        for i, row in enumerate(rows):
            obj = {'rowid': offset + i + 1}  # 1-based rowid
            for j, col in enumerate(columns):
                obj[col] = row[j] if j < len(row) else ''
            formatted_rows.append(obj)

        if rowid_toggle == 'hide':
            for obj in formatted_rows:
                del obj['rowid']

        response_data['rows'] = formatted_rows
    else:
        # lists shape (default)
        response_data['rows'] = rows

    return success_response(response_data)


@app.route('/upload', methods=['POST', 'OPTIONS'])
def upload():
    """Upload a file (CSV, XLS, XLSX) and create a dataset."""
    if request.method == 'OPTIONS':
        return make_json_response({'ok': True}, 200)

    # Check if request is multipart
    if not request.files:
        return error_response("Non-multipart upload", 415)

    # Try to get file from 'file' or 'attachment' field
    uploaded_file = request.files.get('file') or request.files.get('attachment')

    if not uploaded_file:
        return error_response("Missing file field", 400)

    # Read file content
    content = uploaded_file.read()
    filename = uploaded_file.filename or ''

    # Generate dataset ID from file content hash (deterministic)
    dataset_id = hashlib.sha256(content).hexdigest()[:16]

    # Parse the content
    charset = request.form.get('charset')

    try:
        columns, rows = parse_content(content, charset, filename)
    except ValueError as e:
        return error_response(str(e), 400)
    except Exception as e:
        return error_response(f"Failed to parse file: {str(e)}", 400)

    # Store the dataset
    datasets[dataset_id] = {
        'columns': columns,
        'rows': rows,
        'source': f'upload://{filename}'
    }

    return success_response({'endpoint': f'/datasets/{dataset_id}'})


@app.route('/datasets/<dataset_id>/export', methods=['GET', 'OPTIONS'])
def export_dataset(dataset_id):
    """Export a dataset as CSV."""
    if request.method == 'OPTIONS':
        response = Response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return response, 200

    if dataset_id not in datasets:
        return error_response("Dataset not found", 404)

    dataset = datasets[dataset_id]
    columns = dataset['columns']
    rows = dataset['rows']

    # Parse query parameters for filtering, sorting, and pagination
    # (same as /datasets/<id> but only for row selection)

    # Parse and validate filters
    valid_comparators = {'exact', 'contains', 'less', 'greater'}
    filters = {}

    for key in request.args.keys():
        # Skip control parameters (starting with _)
        if key.startswith('_'):
            continue

        # Check if this is a filter parameter (must contain __)
        if '__' not in key:
            continue

        value = request.args.get(key)
        if value is None:
            continue

        # Split on __ to get column and comparator
        parts = key.split('__', 1)
        column = parts[0]
        comparator = parts[1]

        # Validate comparator
        if comparator not in valid_comparators:
            continue

        # Validate column exists
        if column not in columns:
            continue

        filters[(column, comparator)] = value

    # Parse pagination parameters for export
    size = request.args.get('_size')
    offset = request.args.get('_offset', '0')

    if size is not None:
        try:
            size = int(size)
            if size <= 0:
                return error_response("_size must be a positive integer", 400)
        except (ValueError, TypeError):
            return error_response("_size must be a positive integer", 400)
    else:
        size = None  # No limit

    try:
        offset = int(offset)
        if offset < 0:
            return error_response("_offset must be a non-negative integer", 400)
    except (ValueError, TypeError):
        return error_response("_offset must be a non-negative integer", 400)

    # Parse sorting parameters
    sort = request.args.get('_sort')
    sort_desc = request.args.get('_sort_desc')

    if sort is not None:
        if sort == '' or sort not in columns:
            sort = None

    if sort_desc is not None:
        if sort_desc == '' or sort_desc not in columns:
            sort_desc = None

    # Apply filtering before sorting
    if filters:
        filtered_rows = []
        for row in rows:
            match = True
            for (column, comparator), filter_value in filters.items():
                col_index = columns.index(column)
                cell_value = row[col_index] if col_index < len(row) else ''

                if comparator == 'exact':
                    if str(cell_value) != filter_value:
                        match = False
                        break
                elif comparator == 'contains':
                    if filter_value not in str(cell_value):
                        match = False
                        break
                elif comparator == 'less':
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num < filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        match = False
                        break
                elif comparator == 'greater':
                    try:
                        cell_num = float(cell_value)
                        filter_num = float(filter_value)
                        if not (cell_num > filter_num):
                            match = False
                            break
                    except (ValueError, TypeError):
                        match = False
                        break

            if match:
                filtered_rows.append(row)
        rows = filtered_rows

    # Apply sorting (stable sort)
    if sort_desc is not None:
        col_index = columns.index(sort_desc)
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''), reverse=True)
        rows = [r[1] for r in indexed_rows]
    elif sort is not None:
        col_index = columns.index(sort)
        indexed_rows = list(enumerate(rows))
        indexed_rows.sort(key=lambda x: str(x[1][col_index] if col_index < len(x[1]) else ''))
        rows = [r[1] for r in indexed_rows]

    # Apply pagination
    if offset >= len(rows):
        rows = []
    elif size is not None:
        rows = rows[offset:offset + size]
    else:
        rows = rows[offset:]

    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)

    csv_content = output.getvalue()

    # Create response with appropriate headers
    response = Response(
        csv_content,
        mimetype='text/csv',
        headers={
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename="{dataset_id}.csv"',
            'Access-Control-Allow-Origin': '*'
        }
    )

    return response, 200


@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors."""
    return error_response("Not found", 404)


@app.errorhandler(500)
def internal_error(e):
    """Handle 500 errors."""
    return error_response("Internal server error", 500)


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description='datagate - CSV to JSON conversion service')
    subparsers = parser.add_subparsers(dest='command')

    start_parser = subparsers.add_parser('start', help='Start the server')
    start_parser.add_argument('--port', type=int, default=8001, help='Port to listen on (default: 8001)')
    start_parser.add_argument('--address', type=str, default='127.0.0.1', help='Address to bind to (default: 127.0.0.1)')

    args = parser.parse_args()

    if args.command == 'start':
        print(f"Starting datagate on {args.address}:{args.port}")
        app.run(host=args.address, port=args.port)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
